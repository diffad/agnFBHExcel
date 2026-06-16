# -*- coding: utf-8 -*-
"""Aufbau der Excel-Arbeitsmappe zur ueberschlaegigen Auslegung von Fussbodenheizungen."""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.cell.cell import MergedCell
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.chart import LineChart, Reference
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont

CHANGELOG = [
    ("0.1", "2026-06-07", [
        "Ersterstellung mit den Blättern Grundeinstellungen, Raumliste, Übersicht und Methodik.",
        "Heizleistung über die logarithmische Übertemperatur (q = η · K_H · ΔθH).",
        "Druckverlustberechnung nach Darcy-Weisbach.",
        "Editierbare Rohrbibliothek und Verlegeabstand-Faktoren.",
        "Prüfungen: Oberflächentemperatur, Heizlastdeckung, Kreislänge, Druckverlust.",
    ]),
    ("0.2", "2026-06-07", [
        "agn-Logo eingebunden und eigenes Changelog-Blatt ergänzt.",
        "Zuschlag Einzelwiderstände (5 %) und fester Verteiler-Aufschlag (500 Pa je Kreis).",
        "Einheiten als Zellformat; Projektnummer/Projektname erscheinen im Ausdruck.",
        "Richtwert-Tabellen für R-Werte und Verlegeabstand-Faktoren; weitere Musterräume.",
        "Druckaufbereitung für DIN A4.",
    ]),
    ("0.3", "2026-06-07", [
        "Zonentypen (Aufenthalt/Rand/Bad) mit eigener zulässiger Oberflächentemperatur.",
        "Einheitlicher Kopf: Titel oben links, agn-Logo oben rechts, Projektkopf darunter.",
        "Oberflächentemperatur θF und Volumenstrom eingeblendet.",
        "Neue Spalten 'spez. Druckverlust [Pa/m]' und Eingabespalte 'Verteiler (Anbindung)'.",
        "Warnung (rot), wenn die aktivierbare Fläche größer als die Raumfläche ist.",
    ]),
    ("0.4", "2026-06-07", [
        "Status-Spalten entfernt – Bewertung erfolgt jetzt direkt als Zellfärbung (grün/rot).",
        "Methodik um die Formelzeichen-Legende erweitert.",
        "Heizkreisverteiler (HKV) als erste Spalte der Auslegung; Standardwert max. Druckverlust 15.000 Pa.",
        "Versionsschema 0.x eingeführt (1.0 erst nach Test); interner Autor (dh).",
    ]),
    ("0.5", "2026-06-07", [
        "Versionsnummer und Autor werden nur noch im Changelog angezeigt.",
        "Editierbares Bearbeiter-Kürzel rechts oben unter dem Logo (Grundeinstellungen, Auslegung, Kontrolle, HKV).",
    ]),
    ("0.6", "2026-06-07", [
        "Bugfix: Zellfärbung (bedingte Formatierung) wird jetzt auch in Excel angezeigt – dxf-Füllfarbe mit voller Deckkraft.",
    ]),
    ("0.7", "2026-06-07", [
        "Deckung-%-Spalte färbt nur noch gefüllte Zellen (leere bleiben farblos).",
        "Neues Blatt 'Verifikation'; neue Spalte 'Strömungsbereich' (laminar/turbulent).",
        "Alle Blätter im A3-Querformat; einheitlicher Kopf je Blatt.",
    ]),
    ("0.8", "2026-06-07", [
        "Grundeinstellungen für A3 quer optimiert (Einstellungen links, Tabellen rechts).",
        "Auslegung passt auf eine Seite breit; Korrekturfaktor Heizleistung zur Kalibrierung ergänzt.",
        "Wärmeabgabe nach unten berücksichtigt (Verlust + Gesamtleistung).",
        "Blätter umbenannt: Raumliste → Auslegung, Übersicht → Kontrolle; zusätzliche Kontrollen.",
    ]),
    ("0.9", "2026-06-07", [
        "Neues Blatt 'Heizkreisverteiler' (Leistung, max. Druckverlust, Volumenstrom je HKV).",
        "Spaltenüberschriften der Auslegung dreizeilig (Name / Formelzeichen / Einheit).",
        "Korrekturfaktor f wird auf 'Verifikation' eingegeben (mit Vergleichsdiagramm).",
        "Geschwindigkeits-Grenzwerte v_min/v_max als Grundeinstellung.",
    ]),
    ("0.10", "2026-06-08", [
        "Heizkreisverteiler-Liste und Verifikation ohne dynamische Array-Funktionen → läuft in Excel UND LibreOffice.",
        "Verifikation mit Liniendiagramm; Konstanten/Bibliotheken in eigenes Blatt ausgelagert.",
        "Gewähltes Rohrsystem mit Außen-/Innendurchmesser, Wandstärke und Rauheit im Druckbereich.",
    ]),
    ("0.11", "2026-06-08", [
        "Bugfixes: Hinweistext wurde als Formel interpretiert (Fehler 501); MAXIFS ersetzt (#NAME? in LibreOffice).",
        "Verifikation: Korrekturfaktor prominent oben; Hersteller-q direkt neben q berechnet.",
        "Konstanten: alle Tabellen untereinander, mit Leerzeilen für weitere Einträge.",
        "Methodik: Variablen zuerst, dann Formeln mit Erklärung je Schritt; weitere Musterräume (Verwaltungsbau).",
    ]),
    ("0.12", "2026-06-08", [
        "Grundeinstellungen: Bezeichnungen ausgeschrieben (keine Abkürzungen).",
        "Neue Spalte 'spez. Heizlast [W/m²]' (Heizlast bezogen auf die Raumfläche).",
        "Massenstrom/Geschwindigkeit/Druckverlust auf Basis min(Leistung; Heizlast) + Verlust nach unten (realer Betriebspunkt).",
        "Heizkreisverteiler zusätzlich mit Massenstrom [kg/h]; Zonen-Kürzel (AZ/RZ/BAD).",
        "Kapazität 200 Räume / 50 Verteiler; Auslegung A3 quer, übrige Blätter A4 hoch.",
    ]),
    ("0.13", "2026-06-08", [
        "Strömungsbereich dreistufig: laminar (Re < 2300), Übergangsbereich (2300–4000), turbulent (Re > 4000).",
        "Spalte 'Strömungsbereich' ausgeblendet (für die Auslegung nicht relevant) und ohne Färbung.",
        "Auslegung schlanker: weitere Zwischenspalten standardmäßig ausgeblendet (jederzeit einblendbar).",
        "Verifikation: Korrekturfaktor kleiner dargestellt (passt nun vollständig in die Zelle).",
        "Changelog wieder ausführlich.",
    ]),
    ("0.14", "2026-06-09", [
        "Deckblatt als erstes Blatt: Übersichtsdiagramm Eingaben → Verarbeitung → Ergebnisse (inkl. Vereinfachungen).",
    ]),
    ("0.15", "2026-06-09", [
        "Deckblatt-Diagramm aus Excel-Zellen statt Bild – direkt in Excel bearbeitbar.",
        "Auslegung: Beispielräume mit gemischter Heizlast (≤ 50 W/m²); Formelzeichen in Zeile 5 ergänzt (Flächen u. a.).",
        "Heizkreisverteiler: zusätzliche Spalte Spreizung [K].",
    ]),
    ("0.16", "2026-06-09", [
        "Alle Blätter im A4-Querformat (nur Auslegung A3 quer); einheitlichere Schriftgrößen/Skalierung.",
        "Grundeinstellungen zweispaltig (Einstellwerte teilweise nebeneinander).",
        "Verifikation: Diagramm rechts neben den Tabellen.",
        "Methodik: Formelzeichen links, Berechnungsschritte in zwei Spalten rechts.",
        "Konstanten: Tabellen teilweise nebeneinander.",
        "Auslegung: Spalte Zone mittig/schmaler, Raum-Nr. breiter; Verifikation-Korrekturfaktor vertikal mittig.",
    ]),
    ("0.17", "2026-06-13", [
        "Neues Blatt 'Anleitung' (2. Blatt): Aufbau der Mappe, Schritt-für-Schritt, Farblegende.",
        "Anleitung enthält den Workflow zum Übernehmen der Räume aus Revit (Copy & Paste, Schedule, Dynamo/Power Query).",
    ]),
    ("0.18", "2026-06-13", [
        "Auslegung: Eingabespalten neu sortiert – zuerst die Revit-Spalten (A HKV, B Raum-Nr., C Bezeichnung, D Raumfläche, E Raumtemperatur, F Heizlast), dann G spez. Heizlast (berechnet), danach die manuellen Felder (H aktiv. Fläche, I R-Wert, J Verlegeabstand, K Anz. Kreise, L Zuleitung, M Zone).",
        "Anleitung: Excel-Export aus Revit ist jetzt der Standard-Workflow (direktes Kopieren funktioniert nicht); Spaltenliste an neue Reihenfolge angepasst.",
    ]),
    ("0.19", "2026-06-13", [
        "Grundeinstellungen: neuer Grenzwert 'Maximaler Volumenstrom je Kreis' (V̇_max, Default 150 l/h); Abschnitt 'Strömungsgeschwindigkeit' → 'Strömungsgrenzwerte'.",
        "Auslegung: Volumenstrom je Kreis wird gegen V̇_max geprüft und per Ampel eingefärbt (grün ≤ Grenze, rot darüber); zusätzliche Warnung im Blatt 'Kontrolle'.",
        "Anleitung: durchgehende Schritt-für-Schritt-Anleitung mit integriertem Revit-Export (vorgefertigte Bauteilliste muss nicht erst angelegt werden); redundanter Abschnitt 'Räume aus Revit' und der Hinweis zum direkten Kopieren entfernt.",
        "Deckblatt: Pfeil-Spalten breiter und Pfeilgröße angepasst (optisch sauberer).",
    ]),
    ("0.20", "2026-06-13", [
        "Grundeinstellungen: Spaltenbreiten optimiert und lange Beschriftungen/Hinweise gekürzt – Texte werden in LibreOffice nicht mehr abgeschnitten.",
        "Verifikation: Untertitel und Hinweise vereinfacht (Modellvergleich-Erläuterung entfernt); klarer Hinweis, dass der Korrekturfaktor f VOR der Raumauslegung kalibriert wird; Diagramm etwas größer.",
        "Anleitung: Schritt-für-Schritt neu geordnet – Verifikation/Kalibrierung des Korrekturfaktors jetzt VOR der Raumauslegung; neuer Abschnitt 'Planungstechnische Grundlagen'; Hinweis zur vorgefertigten Revit-Liste entfernt.",
    ]),
    ("0.21", "2026-06-13", [
        "Auslegung jetzt auf A4-Querformat (vorher A3); sichtbare Spalten schmaler, damit das Blatt auf A4 passt.",
        "Deckblatt: Titel 'Fußbodenheizung – Auslegung' (Wort 'überschlägige' entfernt); Vereinfachungen bleiben dokumentiert.",
        "Anleitung: Abschnitt 'Planungstechnische Grundlagen' wieder entfernt.",
    ]),
    ("0.22", "2026-06-13", [
        "Anleitung: neue Ampel-Legende (warum ist eine Zelle rot?) und Hinweis zur Auslegung mehrerer Verteiler/Geschosse; Blatt nach hinten verschoben (jetzt vor der Methodik).",
        "Deckblatt: Druck-Hinweis ergänzt – nur die relevanten Blätter (Deckblatt, Grundeinstellungen, Auslegung, Kontrolle, HKV, Anleitung) gemeinsam als 'Aktive Blätter' drucken.",
        "Formelzeichen mit echten Tiefstellungen dargestellt (z. B. q_HL, R_λ,B, V̇_max) – kein Unterstrich mehr in den Kopfzeilen, Grundeinstellungen und der Methodik-Variablenliste.",
    ]),
    ("0.23", "2026-06-13", [
        "Auslegung: Formelzeichen größer (besser lesbare Tiefstellungen); Kopfzeilen-Umbrüche optimiert – keine Trennung mehr mitten im Wort.",
        "Deckblatt: Druck-Hinweis wieder entfernt.",
    ]),
    ("0.24", "2026-06-14", [
        "Verifikation: Referenzspalte mit echten Werten nach DIN EN 1264 vorbelegt (Nasssystem/Tacker, VL 40 / RL 30 °C, θi 20 °C, R = 0,10 m²K/W); Vergleichsbedingungen entsprechend auf 40/30 °C gesetzt.",
        "Verifikation: Korrekturfaktor f auf 0,85 vorkalibriert (Best-Fit zu den EN-1264-Werten); Spalte C bleibt editierbar für herstellereigene Datenblattwerte (Rehau, Uponor, Buderus, Kermi …).",
    ]),
    ("0.25", "2026-06-14", [
        "Begriff 'Korrekturfaktor' in 'Systemfaktor (EN-1264-Kalibrierung)' umbenannt (Verifikation, Methodik, Deckblatt, Anleitung) – f ist die bewusste Kalibrierung des überschlägigen Modells an EN 1264, kein Fehlerfaktor.",
    ]),
    ("0.26", "2026-06-14", [
        "HKV-Liste: neue Spalte 'angebundene Räume (Raum-Nr.)' – listet je Verteiler automatisch die zugehörigen Raumnummern (TEXTJOIN-Array, Excel & LibreOffice).",
        "Auslegung: Spalten 'Leistung FBH' und 'Über-/Unterdeckung' eingeblendet; Über-/Unterdeckung mit Vorzeichen (+/–) und Ampel (grün = Überdeckung, rot = Unterdeckung).",
        "Grundeinstellungen: Eingabefelder Projekt-Nr./Projektname breiter; Spalte J (Rohrsystem-Werte) breiter.",
        "Verifikation: kein fixer Vergleichspunkt mehr – jede Zeile ist ein eigener Auslegungspunkt (VL/RL, θi, R, Verlegeabstand) mit eigener Abweichung; Diagramm entfernt.",
        "Blattschutz: Zellen mit Formeln sind gesperrt (vor versehentlichem Überschreiben geschützt), Eingabezellen bleiben editierbar; Schutz ohne Passwort.",
    ]),
    ("0.27", "2026-06-14", [
        "Auslegung: 'Über-/Unterdeckung' wieder ausgeblendet; stattdessen wird die 'Leistung FBH' eingefärbt (grün = Heizlast gedeckt, rot = nicht gedeckt). Raum-Nr. breiter; Anzahl Heizkreise mit Einheit 'HK' je Zelle; schmalere Seitenränder für besseren A4-Sitz.",
        "HKV-Liste: Spalte 'angebundene Räume' deutlich breiter (lange/viele Raumnummern werden vollständig dargestellt).",
        "Konstanten: R-Werte-Tabelle der Bodenbeläge erweitert (11 Beläge statt 5) und nach rechts gestellt; Zonen-Tabelle nach links – so kann die R-Werte-Tabelle höher werden.",
    ]),
    ("0.28", "2026-06-15", [
        "Blattschutz: Zeilenhöhe und Spaltenbreite bleiben trotz Schutz änderbar – in der HKV-Liste lässt sich die Zeilenhöhe für lange Raumlisten jetzt wieder anpassen (optimale Höhe).",
        "Auslegung: Spalte 'R-Wert Bodenbelag' → Auswahl jetzt über den Bodenbelag-Namen (Dropdown), der R-Wert wird automatisch aus den Konstanten gezogen – man sieht den Belag statt nur die Zahl.",
        "Auslegung: Bearbeiter-Feld oben rechts als zwei Zellen (Label + eigenes Eingabefeld) und auf sichtbaren Spalten – der Eintrag wird nicht mehr abgeschnitten.",
        "Verifikation: 15 statt 8 Referenzpunkte nach DIN EN 1264 (mehr Prüfstellen: Verlegeabstand 100–300 mm, VL 40/45/50 °C, R 0,00–0,15 m²K/W).",
        "Konstanten: R-Werte-Tabelle auf 17 Beläge erweitert (inkl. 'ohne Belag'); Zonen-Tabelle auf gleiche Höhe wie Verlegeabstand-Tabelle; Leerspalte zwischen Zonen und R-Werten.",
    ]),
    ("0.29", "2026-06-15", [
        "Auslegung: in der Spalte wird wieder direkt der R-Wert ausgewählt (Spalte bleibt schmal). Die Zuordnung R-Wert ↔ Bodenbelag erscheint als Hinweis (Tooltip), sobald man die Zelle anklickt – ohne die Spalte zu verbreitern.",
        "Auslegung: Bearbeiter-Feld wieder aus den Grundeinstellungen referenziert (nur dort einmal eingeben) – daher ungefärbt; auf sichtbaren Spalten, damit der Eintrag nicht abgeschnitten wird.",
    ]),
    ("0.30", "2026-06-16", [
        "Verifikation: Beispielwerte durch 12 reale Herstellerpunkte ersetzt (Rehau 17×2, Estrichüberdeckung s_ü 45 mm; zwei Belag-/Temperaturfälle, Verlegeabstand 50–300 mm).",
        "Verifikation: neue Summenzeile 'Gesamtabweichung (Σ über alle Punkte)' in W/m² und % – zeigt, wie gut der Systemfaktor f insgesamt passt (nahe 0 ⇒ gut kalibriert; bei Rehau ≈ −0,2 %).",
        "Deckblatt: Hinweis-Kasten 'Vereinfachungen' entfernt (die Vereinfachungen sind weiterhin im Blatt 'Methodik' dokumentiert).",
    ]),
    ("0.31", "2026-06-16", [
        "Verifikation: Spaltenköpfe und Hinweis klargestellt – verglichen wird die spez. Heizleistung NACH OBEN (Wärmestrom zur Raumseite); die Wärmeabgabe nach unten ist hier bewusst nicht enthalten.",
        "Deckblatt überarbeitet: Inhalte über die volle A4-quer-Höhe verteilt, fette Überschriften, ausgeschriebene Begriffe (keine Abkürzungen), neuer Block 'Die wichtigsten Berechnungsformeln' mit Formelzeichen-Legende.",
        "Deckblatt/Methodik: Hinweis ergänzt, dass das Tool nur für Bauart A nach DIN EN 1264 (Heizrohre im Estrich) gilt.",
        "Auslegung: Hinweis-/Kommentartext (Tooltip) an der R-Wert-Spalte wieder entfernt (er aktualisiert sich nicht mit den Konstanten).",
    ]),
    ("0.32", "2026-06-16", [
        "Deckblatt neu strukturiert: jede Information in einer eigenen Zelle (keine Mehrzeilen-Blöcke mehr, keine Fehl-Umbrüche); die wichtigsten Formeln stehen jetzt direkt in der Spalte 'BERECHNUNG'.",
        "Deckblatt: Bauart-A-Hinweis entfernt; steht stattdessen bei den Grundeinstellungen unter 'Gewähltes Rohrsystem'.",
        "Verifikation: Systemfaktor-Feld vergrößert (Spalte breiter, Beschriftung etwas kleiner) – der Wert ist jetzt vollständig lesbar.",
    ]),
]
VERSION = CHANGELOG[-1][0]
AUTHOR = "dh"
VERTXT = f"Version {VERSION} · {AUTHOR}"
LOGO = r"C:\Users\derDi\FBH_Auslegung_Excel\agn-logo.png"

FONT = "Arial"
NAME_ROW, SYM_ROW, UNIT_ROW = 4, 5, 6
R0, N_ROWS = 7, 200
R1 = R0 + N_ROWS - 1   # 206
EMU = 9525
BLUE, BLACK, WHITE, GREY, NAVY = "0000FF", "000000", "FFFFFF", "808080", "1F4E78"
HDR_FILL = PatternFill("solid", fgColor=NAVY)
SUB_FILL = PatternFill("solid", fgColor="D9E1F2")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
ACCENT_FILL = PatternFill("solid", fgColor="FFE699")
RED_FILL = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")
GREEN_FILL = PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid")
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
CEN = Alignment(horizontal="center")
LEFT = Alignment(horizontal="left")

def f(bold=False, color=BLACK, size=10, italic=False):
    return Font(name=FONT, bold=bold, color=color, size=size, italic=italic)

def fz(text, color=BLACK, bold=False, size=10, italic=False):
    """Formelzeichen mit echten Tiefstellungen: nach '_' folgt bis zum nächsten
    Leerzeichen ein tiefgestelltes Token (z. B. q_HL → q mit tiefem HL)."""
    if "_" not in text:
        return text
    sf = InlineFont(rFont=FONT, vertAlign="subscript", b=bold, i=italic, sz=size, color=color)
    parts, cur, i, n = [], "", 0, len(text)
    while i < n:
        if text[i] == "_" and i + 1 < n and not text[i + 1].isspace():
            if cur:
                parts.append(cur); cur = ""
            i += 1; tok = ""
            while i < n and not text[i].isspace():
                tok += text[i]; i += 1
            parts.append(TextBlock(sf, tok))
        else:
            cur += text[i]; i += 1
    if cur:
        parts.append(cur)
    return CellRichText(parts)

def col_px(ws, idx):
    dim = ws.column_dimensions[get_column_letter(idx)]
    if dim.hidden: return 0
    w = dim.width if dim.width else 8.43
    return int(round(w * 7)) + 5

def add_logo_corner(ws, last_col, height=40):
    if not os.path.exists(LOGO): return
    img = XLImage(LOGO)
    ratio = img.width / img.height
    img.height = height; img.width = int(height * ratio)
    right_edge = sum(col_px(ws, i) for i in range(1, last_col + 1))
    left = max(0, right_edge - img.width - 4)
    acc, c = 0, 1
    while c <= last_col:
        cw = col_px(ws, c)
        if cw and acc + cw > left: break
        acc += cw; c += 1
    marker = AnchorMarker(col=c - 1, colOff=(left - acc) * EMU, row=0, rowOff=1 * EMU)
    img.anchor = OneCellAnchor(_from=marker, ext=XDRPositiveSize2D(cx=img.width * EMU, cy=img.height * EMU))
    ws.add_image(img)

def disp_header(ws, title, last_col, bearb_cell=None, bearb_merge=None, project=True):
    ws.row_dimensions[1].height = 34
    t = ws["A1"]; t.value = title
    t.font = Font(name=FONT, bold=True, color=NAVY, size=15); t.alignment = Alignment(horizontal="left", vertical="center")
    add_logo_corner(ws, last_col, height=40)
    if project:
        a2 = ws["A2"]; a2.value = f'="Projekt-Nr.: "&{gPNr}'; a2.font = f(bold=True, color=NAVY, size=11); a2.alignment = LEFT
        a3 = ws["A3"]; a3.value = f'="Projekt: "&{gPName}'; a3.font = f(bold=True, color=NAVY, size=11); a3.alignment = LEFT
    if bearb_cell:
        if bearb_merge: ws.merge_cells(bearb_merge)
        b = ws[bearb_cell]; b.value = f'="Bearbeiter: "&{gBearb}'
        b.font = f(italic=True, color=GREY, size=9); b.alignment = Alignment(horizontal="right", vertical="center")

def setup_print(ws, area, titles=None, orientation="landscape", paper=9):
    # paper: 9 = A4, 8 = A3. Auf eine Seite breit skaliert (fit-to-width).
    ws.page_setup.orientation = orientation
    ws.page_setup.paperSize = paper
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_area = area
    if titles: ws.print_title_rows = titles
    ws.page_margins.left = ws.page_margins.right = 0.4
    ws.page_margins.top = ws.page_margins.bottom = 0.5
    ws.oddFooter.left.text = "FBH überschlägige Auslegung"
    ws.oddFooter.center.text = "Seite &P von &N"
    ws.oddFooter.right.text = "Stand &D"

def mini_header(ws, row, col, text):
    c = ws.cell(row=row, column=col, value=text)
    c.font = f(bold=True, color=WHITE); c.fill = HDR_FILL
    c.alignment = Alignment(wrap_text=True, horizontal="center"); c.border = BORDER

def two_col_table(ws, title_row, c0, title, hdr1, hdr2, rows, fmt2, n_empty=0):
    ws.cell(row=title_row, column=c0, value=title).font = f(bold=True, color=NAVY)
    ws.merge_cells(start_row=title_row, start_column=c0, end_row=title_row, end_column=c0 + 1)
    mini_header(ws, title_row + 1, c0, hdr1); mini_header(ws, title_row + 1, c0 + 1, hdr2)
    top = title_row + 2
    total = len(rows) + n_empty
    for i in range(total):
        r = top + i
        a = ws.cell(row=r, column=c0); b = ws.cell(row=r, column=c0 + 1)
        a.font = f(color=BLUE); a.fill = INPUT_FILL; a.border = BORDER
        b.font = f(color=BLUE); b.fill = INPUT_FILL; b.border = BORDER; b.number_format = fmt2; b.alignment = CEN
        if i < len(rows):
            a.value = rows[i][0]; b.value = rows[i][1]
            a.alignment = Alignment(horizontal="left" if isinstance(rows[i][0], str) else "center")
        else:
            a.alignment = LEFT
    return top, top + total - 1

wb = Workbook()

# ---- Referenzen auf KONSTANTEN ----
K = "'Konstanten'!"
VA_RANGE = f"{K}$A$6:$B$13"
RW_VAL_LIST = f"{K}$I$6:$I$25"    # Dropdown: R-Wert (Spalte I); Belag-Name steht daneben (Spalte H)
ZONE_RANGE = f"{K}$E$6:$F$13"; ZONE_LIST = f"{K}$E$6:$E$13"   # Zonen links (D/E/F): Kürzel -> θF,max
PIPE_RANGE = f"{K}$A$18:$E$25"; PIPE_LIST = f"{K}$A$18:$A$25"

# Bodenbeläge (Name, R-Wert) – Quelle für die Konstanten-Tabelle UND den Dropdown-Hinweis
BELAEGE = [
    ("ohne Belag (roh)", 0.000), ("Keramik / Feinsteinzeug", 0.010), ("Fliesen 8 mm", 0.012),
    ("Naturstein / Marmor", 0.015), ("PVC / Vinyl (verklebt)", 0.020), ("Designboden (Klick)", 0.040),
    ("Linoleum 2,5 mm", 0.050), ("Klick-Vinyl (schwimmend)", 0.055), ("Laminat + Trittschall", 0.060),
    ("Parkett / Fertigparkett", 0.075), ("Mehrschichtparkett 14 mm", 0.090), ("Kork", 0.100),
    ("Teppich (dünn)", 0.100), ("Holzdielen", 0.110), ("Massivholzdielen 20 mm", 0.130),
    ("Teppich (mittel)", 0.140), ("Teppich (dick)", 0.170),
]

# =====================================================================
#  Blatt 1: GRUNDEINSTELLUNGEN  (Bezeichnungen ausgeschrieben)
# =====================================================================
g = wb.active
g.title = "Grundeinstellungen"
g.sheet_view.showGridLines = False
for col, w in (("A", 23), ("B", 15), ("C", 11), ("D", 8), ("E", 17), ("F", 7), ("G", 2),
               ("H", 23), ("I", 15), ("J", 14), ("K", 8), ("L", 17), ("M", 7)):
    g.column_dimensions[col].width = w
g.row_dimensions[1].height = 34
g["A1"].value = "Globale Grundeinstellungen – Fußbodenheizung"
g["A1"].font = Font(name=FONT, bold=True, color=NAVY, size=15); g["A1"].alignment = Alignment(horizontal="left", vertical="center")
add_logo_corner(g, 13, height=40)
for row, lab in ((2, "Projekt-Nr."), (3, "Projektname")):
    g.cell(row=row, column=1, value=lab).font = f(bold=True)
    g.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    v = g.cell(row=row, column=3, value=""); v.font = f(bold=True, color=BLUE); v.fill = INPUT_FILL; v.border = BORDER; v.alignment = LEFT
    g.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
g.cell(row=2, column=12, value="Bearbeiter:").font = f(italic=True, color=GREY, size=9)
g["L2"].alignment = Alignment(horizontal="right")
bf = g.cell(row=2, column=13, value=""); bf.font = f(bold=True, color=BLUE); bf.fill = INPUT_FILL; bf.border = BORDER; bf.alignment = CEN

def section(row, text, cL):
    for col in range(cL, cL + 6): g.cell(row=row, column=col).fill = HDR_FILL
    g.cell(row=row, column=cL, value=text).font = f(bold=True, color=WHITE)
def param(row, label, value, unit, note, fmt, cL):
    g.cell(row=row, column=cL, value=fz(label)).font = f()
    g.merge_cells(start_row=row, start_column=cL, end_row=row, end_column=cL + 1)
    v = g.cell(row=row, column=cL + 2, value=value); v.font = f(bold=True, color=BLUE); v.alignment = CEN; v.border = BORDER; v.fill = INPUT_FILL
    if fmt: v.number_format = fmt
    g.cell(row=row, column=cL + 3, value=unit).font = f(color=GREY)
    g.cell(row=row, column=cL + 4, value=note).font = f(italic=True, color=GREY, size=9)
    g.merge_cells(start_row=row, start_column=cL + 4, end_row=row, end_column=cL + 5)
def calcrow(row, label, formula, unit, fmt, cL, note=""):
    g.cell(row=row, column=cL, value=fz(label)).font = f(); g.merge_cells(start_row=row, start_column=cL, end_row=row, end_column=cL + 1)
    v = g.cell(row=row, column=cL + 2, value=formula); v.font = f(bold=True); v.alignment = CEN; v.border = BORDER; v.number_format = fmt
    g.cell(row=row, column=cL + 3, value=unit).font = f(color=GREY)
    if note:
        g.cell(row=row, column=cL + 4, value=note).font = f(italic=True, color=GREY, size=9); g.merge_cells(start_row=row, start_column=cL + 4, end_row=row, end_column=cL + 5)

# ---- linke Spalte (Werte in C) ----
section(5, "Heizmedium / Temperaturen", 1)
param(6, "Vorlauftemperatur  θV", 35, "°C", "Auslegungs-Vorlauf", '0.0" °C"', 1)
param(7, "Rücklauftemperatur  θR", 28, "°C", "Auslegungs-Rücklauf", '0.0" °C"', 1)
calcrow(8, "Spreizung  θV − θR", "=C6-C7", "K", '0.0" K"', 1, "berechnet")
section(10, "Grenzwerte", 1)
param(11, "Maximale zulässige Kreislänge", 120, "m", "Obergrenze je Heizkreis", '0" m"', 1)
param(12, "Maximaler Druckverlust je Kreis", 15000, "Pa", "Warnschwelle", '#,##0" Pa"', 1)
section(14, "Stoffwerte Heizmedium (Wasser ~30 °C)", 1)
param(15, "Dichte  ρ", 995.7, "kg/m³", "Dichte des Mediums", '0.0', 1)
param(16, "spezifische Wärmekapazität  c_p", 4180, "J/(kg·K)", "spez. Wärmekapazität", '#,##0', 1)
param(17, "kinematische Viskosität  ν", 0.000000801, "m²/s", "für die Reynolds-Zahl", '0.00E+00', 1)
section(19, "Strömungsgrenzwerte", 1)
param(20, "Minimale Geschwindigkeit  v_min", 0.10, "m/s", "untere empfohlene Grenze", '0.00" m/s"', 1)
param(21, "Maximale Geschwindigkeit  v_max", 0.50, "m/s", "obere empfohlene Grenze", '0.00" m/s"', 1)
param(22, "Max. Volumenstrom je Kreis  V̇_max", 150, "l/h", "Warnschwelle je Heizkreis", '0" l/h"', 1)
section(23, "Druckverlust-Zuschläge", 1)
param(24, "Zuschlag Einzelwiderstände", 0.05, "-", "Formstücke (5 %)", '0%', 1)
param(25, "Aufschlag Verteiler je Kreis", 500, "Pa", "fester Wert je Verteiler", '#,##0" Pa"', 1)

# ---- rechte Spalte (Werte in J) ----
section(5, "Wärmeübergang / Fußbodenaufbau", 8)
param(6, "Wärmeübergang Oberfläche  α", 10.8, "W/m²K", "Boden → Raum (EN 1264)", '0.0', 8)
param(7, "Estrichüberdeckung über Rohr  s_ü", 0.045, "m", "Estrich über den Rohren", '0.000', 8)
param(8, "Wärmeleitfähigkeit Estrich  λ_E", 1.2, "W/mK", "Zementestrich typ. 1,2", '0.00', 8)
section(10, "Wärmeabgabe nach unten (Dämmung, global)", 8)
param(11, "Wärmedurchlasswiderstand  R_u", 2.0, "m²K/W", "Dämmung unter den Rohren", '0.00', 8)
param(12, "Temperatur unter der Decke  θ_u", 10, "°C", "Raum/Erdreich unter FBH", '0.0" °C"', 8)
calcrow(13, "Wärmestrom nach unten  q_u", "=((C6+C7)/2-J12)/J11", "W/m²", '0.0" W/m²"', 8, "Mittel θm = (θV+θR)/2")
section(15, "Rohrsystem-Auswahl (Werte aus Konstanten)", 8)
g.cell(row=16, column=8, value="Gewähltes Rohrsystem").font = f(); g.merge_cells("H16:I16")
v = g.cell(row=16, column=10, value="PE-Xa 17x2"); v.font = f(bold=True, color=BLUE); v.fill = INPUT_FILL; v.alignment = CEN; v.border = BORDER
g.merge_cells("J16:K16")
g.cell(row=16, column=12, value="Dropdown").font = f(italic=True, color=GREY, size=9); g.merge_cells("L16:M16")
calcrow(17, "→ Außendurchmesser  da", f"=VLOOKUP($J$16,{PIPE_RANGE},2,FALSE)", "mm", '0.0" mm"', 8, "aus Rohrbibliothek")
calcrow(18, "→ Wandstärke  s", f"=VLOOKUP($J$16,{PIPE_RANGE},3,FALSE)", "mm", '0.0" mm"', 8, "aus Rohrbibliothek")
calcrow(19, "→ Innendurchmesser  di", f"=VLOOKUP($J$16,{PIPE_RANGE},4,FALSE)", "mm", '0.0" mm"', 8, "di = da − 2·s")
calcrow(20, "→ Rauheit  k", f"=VLOOKUP($J$16,{PIPE_RANGE},5,FALSE)", "mm", '0.000" mm"', 8, "aus Rohrbibliothek")
calcrow(21, "→ Innendurchmesser  di", "=J19/1000", "m", '0.0000" m"', 8, "für die Berechnung")
calcrow(22, "→ Innenquerschnitt  A_i", "=PI()/4*J21^2", "m²", '0.000000" m²"', 8, "für die Berechnung")
gilt = g.cell(row=23, column=8, value="Gilt für Bauart A nach DIN EN 1264 (Heizrohre im Estrich).")
gilt.font = f(italic=True, color=GREY, size=9); g.merge_cells("H23:M23")
dv_pipe = DataValidation(type="list", formula1=f"={PIPE_LIST}", allow_blank=False)
g.add_data_validation(dv_pipe); dv_pipe.add(g["J16"])
setup_print(g, "A1:M25")

# ---- Globale Referenzen ----
G = "'Grundeinstellungen'!"
gV, gR = f"{G}$C$6", f"{G}$C$7"
gMax, gWarn = f"{G}$C$11", f"{G}$C$12"
grho, gcp, gnu = f"{G}$C$15", f"{G}$C$16", f"{G}$C$17"
gvmin, gvmax = f"{G}$C$20", f"{G}$C$21"
gVdot = f"{G}$C$22"
gzus, gVarm = f"{G}$C$24", f"{G}$C$25"
galp, gsu, glam = f"{G}$J$6", f"{G}$J$7", f"{G}$J$8"
gqdown = f"{G}$J$13"
gdim, gAi = f"{G}$J$21", f"{G}$J$22"
gPNr, gPName, gBearb = f"{G}$C$2", f"{G}$C$3", f"{G}$M$2"
gfkorr = "'Verifikation'!$D$5"

# =====================================================================
#  Blatt 2: AUSLEGUNG
# =====================================================================
rl = wb.create_sheet("Auslegung")
rl.sheet_view.showGridLines = False
# Überschriften vom Nutzer optimiert (Wortrennungen beibehalten); Formelzeichen ergänzt.
columns = [
    ("Heizkreis-\nverteiler", "HKV", "", 16, "text", BLUE),            # A
    ("Raum-Nr.", "", "", 10, "text", BLUE),                             # B
    ("Raum-\nbezeichnung", "", "", 18, "text", BLUE),                  # C
    ("Raum-fläche", "A_R", "[m²]", 10, '0.0" m²"', BLUE),              # D
    ("Raum-\ntemperatur", "θi", "[°C]", 11, '0.0" °C"', BLUE),         # E
    ("Heizlast", "Q", "[W]", 10, '#,##0" W"', BLUE),                   # F
    ("spez. Heizlast", "q_HL", "[W/m²]", 11, '0.0" W/m²"', BLACK),     # G  (=Q/Raumfläche)
    ("aktivier-\nbare Fläche", "A_F", "[m²]", 11, '0.0" m²"', BLUE),   # H
    ("R-Wert\nBodenbelag", "R_λ,B", "[m²·K/W]", 12, '0.000', BLUE),    # I (Dropdown R-Wert; Belag-Hinweis im Tooltip)
    ("Verlege–abstand", "VA", "[mm]", 11, '0" mm"', BLUE),             # J
    ("Anz.\nHK", "n", "[-]", 9, '0" HK"', BLUE),                       # K
    ("Zuleitungs-\nlänge", "L_zu", "[m]", 11, '0.0" m"', BLUE),        # L
    ("Zone", "", "", 7, "text", BLUE),                                 # M (Kürzel)
    ("log. Übertemperatur", "ΔθH", "[K]", 12, '0.00" K"', BLACK),      # N
    ("Wärmedurchgang", "K_H", "[W/m²K]", 11, '0.000', BLACK),          # O
    ("Verlegeabstand-Faktor", "η", "[-]", 11, '0.000', BLACK),         # P
    ("spez. Heiz-\nleistung", "q", "[W/m²]", 12, '0.0" W/m²"', BLACK), # Q
    ("Ober-\nflächen-\ntemperatur", "θF", "[°C]", 11, '0.0" °C"', BLACK),  # R
    ("max. Oberflächentemp.", "θF,max", "[°C]", 11, '0.0" °C"', BLACK),# S
    ("Leistung\nFBH", "Q_o", "[W]", 12, '#,##0" W"', BLACK),           # T
    ("Über-/Unter-\ndeckung", "ΔQ", "[W]", 12, '"+"#,##0" W";"-"#,##0" W";0" W"', BLACK),  # U
    ("Deckung", "", "[%]", 10, '0.0%', BLACK),                         # V
    ("Verlust nach unten", "Q_u", "[W]", 11, '#,##0" W"', BLACK),      # W
    ("maßgebl. Leistung", "Q_m", "[W]", 12, '#,##0" W"', BLACK),       # X  (=min(T;Q)+W)
    ("Rohrlänge Fläche", "L_F", "[m]", 11, '0.0" m"', BLACK),          # Y
    ("Rohrlänge Zuleitung", "L_Z", "[m]", 11, '0.0" m"', BLACK),       # Z
    ("Rohrlänge gesamt", "L_ges", "[m]", 11, '0.0" m"', BLACK),        # AA
    ("Rohrlänge je Kreis", "L_K", "[m]", 11, '0.0" m"', BLACK),        # AB
    ("Massenstrom je Kreis", "ṁ", "[kg/h]", 11, '0.0" kg/h"', BLACK),  # AC
    ("Volumen-strom je Kreis", "V̇", "[l/h]", 11, '0.0" l/h"', BLACK), # AD
    ("Geschwin-digkeit", "v", "[m/s]", 11, '0.000" m/s"', BLACK),      # AE
    ("Reynolds-Zahl", "Re", "[-]", 11, '#,##0', BLACK),                # AF
    ("Strömungsbereich", "", "[-]", 12, "text", BLACK),                # AG
    ("Rohrreibungszahl", "λ", "[-]", 11, '0.0000', BLACK),             # AH
    ("Δp Reibung", "Δp_R", "[Pa]", 11, '#,##0" Pa"', BLACK),           # AI
    ("spez. Druckverlust", "R", "[Pa/m]", 12, '#,##0" Pa/m"', BLACK),  # AJ
    ("Druck-\nverlust", "Δp", "[Pa]", 12, '#,##0" Pa"', BLACK),        # AK
]  # A..AK = 37
NCOL = len(columns)
LASTCOL = get_column_letter(NCOL)
INPUT_COLS = [j for j, c in enumerate(columns, start=1) if c[5] == BLUE]
CALC_COLS = [j for j, c in enumerate(columns, start=1) if c[5] == BLACK]
INPUT_STYLED = {7}   # spez. Heizlast: berechnet, aber optisch wie die Eingabespalten
for j, (name, sym, unit, width, fmt, color) in enumerate(columns, start=1):
    rl.column_dimensions[get_column_letter(j)].width = width
    is_inp = (color == BLUE) or (j in INPUT_STYLED)
    fill = SUB_FILL if is_inp else HDR_FILL
    tcol = NAVY if is_inp else WHITE
    for row, txt, sz in ((NAME_ROW, name, 10), (SYM_ROW, sym, 13), (UNIT_ROW, unit, 9)):
        val = fz(txt, color=tcol, bold=True, size=sz) if row == SYM_ROW else txt
        c = rl.cell(row=row, column=j, value=val)
        c.fill = fill; c.font = f(bold=True, color=tcol, size=sz)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); c.border = BORDER
rl.row_dimensions[NAME_ROW].height = 34
rl.row_dimensions[SYM_ROW].height = 20
rl.row_dimensions[UNIT_ROW].height = 14

def F(r):
    return {
        7:  f'=IF($B{r}="","",IFERROR($F{r}/$D{r},""))',
        14: f'=IF($B{r}="","",IFERROR(({gV}-{gR})/LN(({gV}-$E{r})/({gR}-$E{r})),""))',
        15: f'=IF($B{r}="","",IFERROR(1/(1/{galp}+$I{r}+{gsu}/{glam}),""))',
        16: f'=IF($B{r}="","",IFERROR(VLOOKUP($J{r},{VA_RANGE},2,TRUE),""))',
        17: f'=IF(OR($B{r}="",$N{r}="",$O{r}="",$P{r}=""),"",{gfkorr}*$O{r}*$P{r}*$N{r})',
        18: f'=IF(OR($B{r}="",$Q{r}=""),"",$E{r}+$Q{r}/{galp})',
        19: f'=IF($B{r}="","",IFERROR(VLOOKUP($M{r},{ZONE_RANGE},2,FALSE),29))',
        20: f'=IF(OR($B{r}="",$Q{r}=""),"",$Q{r}*$H{r})',
        21: f'=IF(OR($B{r}="",$T{r}=""),"",$T{r}-$F{r})',
        22: f'=IF(OR($B{r}="",$T{r}=""),"",IFERROR($T{r}/$F{r},""))',
        23: f'=IF(OR($B{r}="",$H{r}=""),"",{gqdown}*$H{r})',
        24: f'=IF(OR($B{r}="",$T{r}=""),"",MIN($T{r},$F{r})+$W{r})',
        25: f'=IF($B{r}="","",IFERROR($H{r}/($J{r}/1000),""))',
        26: f'=IF($B{r}="","",IFERROR(2*$L{r}*$K{r},""))',
        27: f'=IF(OR($B{r}="",$Y{r}="",$Z{r}=""),"",$Y{r}+$Z{r})',
        28: f'=IF(OR($B{r}="",$AA{r}=""),"",IFERROR($AA{r}/$K{r},""))',
        29: f'=IF(OR($B{r}="",$X{r}=""),"",IFERROR(($X{r}/$K{r})/({gcp}*({gV}-{gR}))*3600,""))',
        30: f'=IF(OR($B{r}="",$AC{r}=""),"",$AC{r}/{grho}*1000)',
        31: f'=IF(OR($B{r}="",$AC{r}=""),"",IFERROR($AC{r}/3600/{grho}/{gAi},""))',
        32: f'=IF(OR($B{r}="",$AE{r}=""),"",IFERROR($AE{r}*{gdim}/{gnu},""))',
        33: f'=IF(OR($B{r}="",$AF{r}=""),"",IF($AF{r}<2300,"laminar",IF($AF{r}<4000,"Übergangsbereich","turbulent")))',
        34: f'=IF(OR($B{r}="",$AF{r}=""),"",IF($AF{r}<2300,64/$AF{r},0.316/$AF{r}^0.25))',
        35: f'=IF(OR($B{r}="",$AH{r}="",$AB{r}=""),"",IFERROR($AH{r}*($AB{r}/{gdim})*({grho}/2)*$AE{r}^2,""))',
        36: f'=IF(OR($B{r}="",$AI{r}="",$AB{r}=""),"",IFERROR($AI{r}/$AB{r},""))',
        37: f'=IF(OR($B{r}="",$AI{r}=""),"",$AI{r}*(1+{gzus})+{gVarm})',
    }

examples = [   # Heizlast so gewählt, dass spez. Heizlast (Q/Raumfläche) <= 50 W/m², bunte Mischung
    ["HKV EG", "001", "Foyer/Empfang", 40, 20, 1800, 36, 0.00,  100, 3, 6,  "RZ"],   # 45 W/m²
    ["HKV EG", "002", "Großraumbüro",  60, 20, 2400, 55, 0.10,  150, 4, 8,  "AZ"],   # 40 W/m²
    ["HKV EG", "003", "Büro 1",        18, 20,  630, 16, 0.10,  150, 1, 5,  "AZ"],   # 35 W/m²
    ["HKV EG", "004", "Besprechung",   30, 20, 1500, 28, 0.075, 100, 2, 7,  "AZ"],   # 50 W/m²
    ["HKV EG", "005", "Flur",          25, 20,  750, 22, 0.05,  200, 1, 3,  "AZ"],   # 30 W/m²
    ["HKV OG", "101", "WC / Sanitär",  12, 24,  600,  9, 0.01,  100, 1, 10, "BAD"],  # 50 W/m²
    ["HKV OG", "102", "Teeküche",      10, 20,  450,  8, 0.01,  100, 1, 9,  "AZ"],   # 45 W/m²
    ["HKV OG", "103", "Archiv",        20, 18,  500, 18, 0.05,  200, 1, 12, "AZ"],   # 25 W/m²
    ["HKV OG", "104", "Technikraum",   15, 15,  300, 12, 0.00,  250, 1, 14, "AZ"],   # 20 W/m²
]
for idx, r in enumerate(range(R0, R1 + 1)):
    fr = F(r)
    for ji, j in enumerate(INPUT_COLS):
        c = rl.cell(row=r, column=j)
        if idx < len(examples): c.value = examples[idx][ji]
        c.font = f(color=BLUE); c.fill = INPUT_FILL; c.border = BORDER
        fmt = columns[j - 1][4]
        if fmt != "text": c.number_format = fmt
        c.alignment = Alignment(horizontal="left" if j in (1, 2, 3) else "center")
    for j in CALC_COLS:
        c = rl.cell(row=r, column=j, value=fr[j])
        c.font = f(color=BLACK); c.border = BORDER
        if j in INPUT_STYLED: c.fill = INPUT_FILL
        fmt = columns[j - 1][4]
        if fmt != "text": c.number_format = fmt
        c.alignment = CEN

rl.freeze_panes = "C7"
dv_zone = DataValidation(type="list", formula1=f"={ZONE_LIST}", allow_blank=True, showErrorMessage=False)
rl.add_data_validation(dv_zone); dv_zone.add(f"M{R0}:M{R1}")
dv_rw = DataValidation(type="list", formula1=f"={RW_VAL_LIST}", allow_blank=True, showErrorMessage=False)
rl.add_data_validation(dv_rw); dv_rw.add(f"I{R0}:I{R1}")

def cf_pair(col, ok_formula, bad_formula):
    rng = f"{col}{R0}:{col}{R1}"
    rl.conditional_formatting.add(rng, FormulaRule(formula=[bad_formula], fill=RED_FILL))
    rl.conditional_formatting.add(rng, FormulaRule(formula=[ok_formula], fill=GREEN_FILL))
cf_pair("R", f'AND($R{R0}<>"",$R{R0}<=$S{R0})', f'AND($R{R0}<>"",$R{R0}>$S{R0})')
cf_pair("T", f'AND($T{R0}<>"",$F{R0}<>"",$T{R0}>=$F{R0})', f'AND($T{R0}<>"",$F{R0}<>"",$T{R0}<$F{R0})')
cf_pair("V", f'AND($V{R0}<>"",$V{R0}>=1)', f'AND($V{R0}<>"",$V{R0}<1)')
cf_pair("AB", f'AND($AB{R0}<>"",$AB{R0}<={gMax})', f'AND($AB{R0}<>"",$AB{R0}>{gMax})')
cf_pair("AD", f'AND($AD{R0}<>"",$AD{R0}<={gVdot})', f'AND($AD{R0}<>"",$AD{R0}>{gVdot})')
cf_pair("AK", f'AND($AK{R0}<>"",$AK{R0}<={gWarn})', f'AND($AK{R0}<>"",$AK{R0}>{gWarn})')
cf_pair("AE", f'AND($AE{R0}<>"",$AE{R0}>={gvmin},$AE{R0}<={gvmax})',
        f'AND($AE{R0}<>"",OR($AE{R0}<{gvmin},$AE{R0}>{gvmax}))')
rl.conditional_formatting.add(f"H{R0}:H{R1}", FormulaRule(
    formula=[f'AND($H{R0}<>"",$D{R0}<>"",$H{R0}>$D{R0})'], fill=RED_FILL))
# schlanker Standard: Zwischen-/Sekundärspalten ausblenden (jederzeit einblendbar)
for col in ["N", "O", "P", "S", "U", "W", "X", "Y", "Z", "AA", "AC", "AF", "AG", "AH", "AI", "AJ"]:
    rl.column_dimensions[col].hidden = True
# sichtbare Spalten schmaler, damit das Blatt auf A4-Querformat passt
narrow = {"A": 13, "B": 11, "C": 16, "D": 8, "E": 11, "F": 9, "G": 10, "H": 9, "I": 11,
          "J": 9, "K": 8, "L": 11, "M": 6, "Q": 11, "R": 11, "T": 9, "V": 8,
          "AB": 9, "AD": 10, "AE": 10, "AK": 10}
for col, w in narrow.items():
    rl.column_dimensions[col].width = w
rl.row_dimensions[NAME_ROW].height = 46   # mehr Höhe, da schmale Spalten stärker umbrechen
# Bearbeiter oben rechts: aus den Grundeinstellungen referenziert (nicht händisch, daher ungefärbt).
# Breite Merge über sichtbare Spalten (AB..AK), damit der Eintrag nicht abgeschnitten wird.
disp_header(rl, "Auslegung – Fußbodenheizung", NCOL, "AB2", "AB2:AK2")
setup_print(rl, f"A1:{LASTCOL}{R1}", titles="1:6", orientation="landscape", paper=9)
rl.page_margins.left = rl.page_margins.right = 0.2   # schmalere Ränder → mehr Platz auf A4
rl.page_margins.top = rl.page_margins.bottom = 0.3

# =====================================================================
#  Blatt 3: KONTROLLE
# =====================================================================
ov = wb.create_sheet("Kontrolle")
ov.sheet_view.showGridLines = False
for col, w in (("A", 44), ("B", 16), ("C", 11), ("D", 40)):
    ov.column_dimensions[col].width = w
AUS = "'Auslegung'!"
def ovrow(r, label, formula, unit, fmt='#,##0', note=""):
    ov.cell(row=r, column=1, value=label).font = f()
    c = ov.cell(row=r, column=2, value=formula); c.font = f(bold=True); c.number_format = fmt
    c.alignment = CEN; c.border = BORDER
    ov.cell(row=r, column=3, value=unit).font = f(color=GREY)
    if note: ov.cell(row=r, column=4, value=note).font = f(italic=True, color=GREY, size=9)
c = ov.cell(row=5, column=1, value="Summen / Kennzahlen"); c.font = f(bold=True, color=WHITE); c.fill = HDR_FILL
for col in (2, 3, 4): ov.cell(row=5, column=col).fill = HDR_FILL
ovrow(6, "Anzahl Räume", f"=COUNTA({AUS}B{R0}:B{R1})", "-", '0')
ovrow(7, "Gesamte Heizlast", f"=SUM({AUS}F{R0}:F{R1})", "W")
ovrow(8, "Gesamte FBH-Leistung (nach oben)", f"=SUM({AUS}T{R0}:T{R1})", "W")
ovrow(9, "Verlustleistung nach unten", f"=SUM({AUS}W{R0}:W{R1})", "W")
ovrow(10, "Maßgebliche Gesamtleistung (Hydraulik)", f"=SUM({AUS}X{R0}:X{R1})", "W")
ovrow(11, "Deckungsgrad gesamt", "=IFERROR(B8/B7,0)", "%", '0.0%')
ovrow(12, "Gesamte Rohrlänge", f"=SUM({AUS}AA{R0}:AA{R1})", "m")
ovrow(13, "Anzahl Heizkreise gesamt", f"=SUM({AUS}K{R0}:K{R1})", "-", '0')
ovrow(14, "Gesamtmassenstrom (Pumpe)", f"=SUM({AUS}AC{R0}:AC{R1})", "kg/h", '#,##0')
ovrow(15, "Gesamtvolumenstrom (Pumpe)", f"=SUM({AUS}AD{R0}:AD{R1})", "l/h", '#,##0')
ovrow(16, "Max. Druckverlust eines Kreises", f"=IFERROR(MAX({AUS}AK{R0}:AK{R1}),0)", "Pa")
ovrow(17, "Max. Strömungsgeschwindigkeit", f"=IFERROR(MAX({AUS}AE{R0}:AE{R1}),0)", "m/s", '0.000')
c = ov.cell(row=19, column=1, value="Warnungen"); c.font = f(bold=True, color=WHITE); c.fill = HDR_FILL
for col in (2, 3, 4): ov.cell(row=19, column=col).fill = HDR_FILL
ovrow(20, "Räume unterdeckt", f'=COUNTIF({AUS}V{R0}:V{R1},"<1")', "-", '0', "Heizlast nicht gedeckt")
ovrow(21, "Kreise zu lang", f'=COUNTIF({AUS}AB{R0}:AB{R1},">"&{gMax})', "-", '0', "> max. Kreislänge")
ovrow(22, "Druckverlust über Warnschwelle", f'=COUNTIF({AUS}AK{R0}:AK{R1},">"&{gWarn})', "-", '0')
ovrow(23, "Oberflächentemperatur zu hoch", f'=SUMPRODUCT(({AUS}R{R0}:R{R1}<>"")*({AUS}R{R0}:R{R1}>{AUS}S{R0}:S{R1}))', "-", '0')
ovrow(24, "Geschwindigkeit außerhalb v_min–v_max", f'=SUMPRODUCT(({AUS}AE{R0}:AE{R1}<>"")*(({AUS}AE{R0}:AE{R1}<{gvmin})+({AUS}AE{R0}:AE{R1}>{gvmax})))', "-", '0', "Kreise")
ovrow(25, "Volumenstrom je Kreis zu hoch", f'=COUNTIF({AUS}AD{R0}:AD{R1},">"&{gVdot})', "-", '0', "> max. Volumenstrom")
for r in (20, 21, 22, 23, 24, 25):
    ov.conditional_formatting.add(f"B{r}", FormulaRule(formula=[f"$B${r}>0"], fill=RED_FILL))
    ov.conditional_formatting.add(f"B{r}", FormulaRule(formula=[f"$B${r}=0"], fill=GREEN_FILL))
disp_header(ov, "Kontrolle / Auswertung", 4, "D2")
setup_print(ov, "A1:D25")

# =====================================================================
#  Blatt 4: HEIZKREISVERTEILER  (+ Massenstrom)
# =====================================================================
hv = wb.create_sheet("HKV")
hv.sheet_view.showGridLines = False
for col, w in (("A", 20), ("B", 10), ("C", 10), ("D", 17), ("E", 14), ("F", 14), ("G", 16), ("H", 50)):
    hv.column_dimensions[col].width = w
for j, h in enumerate(["Heizkreisverteiler (HKV)", "Anzahl Kreise", "Spreizung [K]", "maßgebl. Leistung [W]",
                       "Massenstrom [kg/h]", "Volumenstrom [l/h]", "max. Druckverlust [Pa]",
                       "angebundene Räume (Raum-Nr.)"], start=1):
    c = hv.cell(row=4, column=j, value=h)
    c.font = f(bold=True, color=WHITE); c.fill = HDR_FILL
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); c.border = BORDER
hv.row_dimensions[4].height = 30
Aaus = f"{AUS}$A${R0}:$A${R1}"
N_HKV = 50
for i in range(N_HKV):
    r = 5 + i
    af = (f'=IFERROR(INDEX({Aaus},MATCH(0,COUNTIF($A$4:$A{r-1},{Aaus})+IF({Aaus}="",1,0),0)),"")')
    cell = hv.cell(row=r, column=1); cell.value = ArrayFormula(f"A{r}", af); cell.font = f(); cell.alignment = LEFT; cell.border = BORDER
    b = hv.cell(row=r, column=2, value=f'=IF($A{r}="","",SUMIF({Aaus},$A{r},{AUS}$K${R0}:$K${R1}))')
    sp = hv.cell(row=r, column=3, value=f"=IF($A{r}=\"\",\"\",{gV}-{gR})")   # globale Spreizung (überall gleich)
    c = hv.cell(row=r, column=4, value=f'=IF($A{r}="","",SUMIF({Aaus},$A{r},{AUS}$X${R0}:$X${R1}))')
    d = hv.cell(row=r, column=5, value=f'=IF($A{r}="","",SUMIF({Aaus},$A{r},{AUS}$AC${R0}:$AC${R1}))')
    e = hv.cell(row=r, column=6, value=f'=IF($A{r}="","",SUMIF({Aaus},$A{r},{AUS}$AD${R0}:$AD${R1}))')
    p = hv.cell(row=r, column=7)
    p.value = ArrayFormula(f"G{r}", f'=IF($A{r}="","",MAX(IF({Aaus}=$A{r},{AUS}$AK${R0}:$AK${R1})))')
    for cc, fmt in ((b, '0'), (sp, '0.0" K"'), (c, '#,##0" W"'), (d, '#,##0" kg/h"'), (e, '#,##0" l/h"'), (p, '#,##0" Pa"')):
        cc.font = f(); cc.alignment = CEN; cc.number_format = fmt; cc.border = BORDER
    rm = hv.cell(row=r, column=8)
    rm.value = ArrayFormula(f"H{r}", f'=IF($A{r}="","",_xlfn.TEXTJOIN(", ",TRUE,IF({Aaus}=$A{r},{AUS}$B${R0}:$B${R1},"")))')
    rm.font = f(); rm.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True); rm.border = BORDER
hv.cell(row=5 + N_HKV + 1, column=1,
        value="Baut sich automatisch aus der Auslegung auf (Array-Formeln, Excel & LibreOffice).").font = f(italic=True, color=GREY, size=9)
disp_header(hv, "Heizkreisverteiler – Übersicht je HKV", 8, "F2", "F2:G2")
setup_print(hv, f"A1:H{5 + N_HKV + 1}")

# =====================================================================
#  Blatt 5: VERIFIKATION
# =====================================================================
vf = wb.create_sheet("Verifikation")
vf.sheet_view.showGridLines = False
for col, w in (("A", 13), ("B", 7), ("C", 7), ("D", 11), ("E", 11), ("F", 13),
               ("G", 13), ("H", 12), ("I", 10), ("J", 9), ("K", 10), ("L", 8)):
    vf.column_dimensions[col].width = w
# Systemfaktor f – eine globale Kalibrierung des Modells an EN 1264
vf.cell(row=5, column=1, value="Systemfaktor (EN 1264)  f =").font = f(bold=True, color=NAVY, size=10)
vf.merge_cells("A5:C5")
for cc in ("A5", "B5", "C5"): vf[cc].fill = ACCENT_FILL
vf["A5"].alignment = Alignment(horizontal="right", vertical="center")
fk = vf.cell(row=5, column=4, value=0.85); fk.font = f(bold=True, color=BLUE, size=11); fk.fill = INPUT_FILL
fk.border = BORDER; fk.alignment = Alignment(horizontal="center", vertical="center"); fk.number_format = '0.000'
vf.cell(row=5, column=5, value="← Best-Fit zu Rehau 17×2 (s_ü 45 mm) / EN 1264").font = f(italic=True, color=GREY, size=9)
vf.row_dimensions[5].height = 22
vf.cell(row=6, column=1, value="Verglichen wird die spezifische Heizleistung NACH OBEN (Wärmestrom zur Raumseite). Die Wärmeabgabe nach unten ist hier bewusst NICHT enthalten – wie in den Hersteller-/EN-1264-Kennwerten.").font = f(italic=True, color=GREY, size=9)
vf.merge_cells("A6:L6")
vhdr = ["Verlege-\nabstand [mm]", "θV\n[°C]", "θR\n[°C]", "θi\n[°C]", "R-Wert\n[m²K/W]",
        "Hersteller q\nnach oben\n[W/m²]", "q berechnet\nnach oben\n[W/m²]", "Abweichung\n[W/m²]", "Abweichung\n[%]",
        "ΔθH\n[K]", "K_H\n[W/m²K]", "η\n[-]"]
VT = 8
for j, h in enumerate(vhdr, start=1):
    inp = j <= 6
    c = vf.cell(row=VT, column=j, value=h)
    c.fill = SUB_FILL if inp else HDR_FILL
    c.font = f(bold=True, color=(NAVY if inp else WHITE))
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); c.border = BORDER
vf.row_dimensions[VT].height = 48
V0 = VT + 1
# Beispiel-Auslegungspunkte: [VA, θV, θR, θi, R, Hersteller-q]
# Reale Herstellerdaten: REHAU 17×2-System, Estrichüberdeckung s_ü = 45 mm (= globaler Default).
vpts = [
    [300, 38, 30, 21, 0.125, 30],
    [250, 38, 30, 21, 0.125, 33],
    [200, 38, 30, 21, 0.125, 36],
    [150, 38, 30, 21, 0.125, 40],
    [100, 38, 30, 21, 0.125, 44],
    [50,  38, 30, 21, 0.125, 49],
    [300, 45, 35, 15, 0.050, 74],
    [250, 45, 35, 15, 0.050, 83],
    [200, 45, 35, 15, 0.050, 94],
    [150, 45, 35, 15, 0.050, 105],
    [100, 45, 35, 15, 0.050, 119],
    [50,  45, 35, 15, 0.050, 135],
]
V1 = V0 + len(vpts) - 1
for i, pt in enumerate(vpts):
    r = V0 + i
    inputs = {1: (pt[0], '0" mm"'), 2: (pt[1], '0.0" °C"'), 3: (pt[2], '0.0" °C"'),
              4: (pt[3], '0.0" °C"'), 5: (pt[4], '0.000'), 6: (pt[5], '0.0" W/m²"')}
    for col, (val, fmt) in inputs.items():
        cc = vf.cell(row=r, column=col, value=val); cc.font = f(color=BLUE); cc.fill = INPUT_FILL
        cc.alignment = CEN; cc.number_format = fmt; cc.border = BORDER
    forms = {
        10: (f'=IF($A{r}="","",IFERROR(($B{r}-$C{r})/LN(($B{r}-$D{r})/($C{r}-$D{r})),""))', '0.00" K"'),
        11: (f'=IF($A{r}="","",IFERROR(1/(1/{galp}+$E{r}+{gsu}/{glam}),""))', '0.000'),
        12: (f'=IF($A{r}="","",IFERROR(VLOOKUP($A{r},{VA_RANGE},2,TRUE),""))', '0.000'),
        7:  (f'=IF($A{r}="","",$D$5*$L{r}*$K{r}*$J{r})', '0.0" W/m²"'),
        8:  (f'=IF(OR($A{r}="",$F{r}=""),"",$F{r}-$G{r})', '0.0" W/m²"'),
        9:  (f'=IF(OR($A{r}="",$F{r}=""),"",IFERROR(($F{r}-$G{r})/$G{r},""))', '0.0%'),
    }
    for col, (formula, fmt) in forms.items():
        cc = vf.cell(row=r, column=col, value=formula); cc.font = f(); cc.alignment = CEN; cc.number_format = fmt; cc.border = BORDER
vf.conditional_formatting.add(f"I{V0}:I{V1}", FormulaRule(formula=[f'AND($I{V0}<>"",ABS($I{V0})>0.1)'], fill=RED_FILL))
vf.conditional_formatting.add(f"I{V0}:I{V1}", FormulaRule(formula=[f'AND($I{V0}<>"",ABS($I{V0})<=0.1)'], fill=GREEN_FILL))
# Summenzeile: Gesamtabweichung über alle Punkte (≈ 0 ⇒ Systemfaktor f gut kalibriert)
VS = V1 + 1
slab = vf.cell(row=VS, column=1, value="Gesamtabweichung (Σ über alle Punkte) – nahe 0 ⇒ f gut kalibriert:")
slab.font = f(bold=True, color=NAVY); slab.alignment = Alignment(horizontal="right", vertical="center")
vf.merge_cells(start_row=VS, start_column=1, end_row=VS, end_column=7)
for cc in range(1, 8): vf.cell(row=VS, column=cc).fill = SUB_FILL
sh = vf.cell(row=VS, column=8, value=f"=SUM($H${V0}:$H${V1})")
sh.font = f(bold=True); sh.number_format = '"+"0.0" W/m²";"-"0.0" W/m²";0" W/m²"'
sh.alignment = CEN; sh.border = BORDER; sh.fill = SUB_FILL
si = vf.cell(row=VS, column=9, value=f'=IFERROR(SUM($H${V0}:$H${V1})/SUM($G${V0}:$G${V1}),"")')
si.font = f(bold=True); si.number_format = '"+"0.0%;"-"0.0%;0.0%'
si.alignment = CEN; si.border = BORDER
vf.conditional_formatting.add(f"I{VS}", FormulaRule(formula=[f'AND($I{VS}<>"",ABS($I{VS})>0.03)'], fill=RED_FILL))
vf.conditional_formatting.add(f"I{VS}", FormulaRule(formula=[f'AND($I{VS}<>"",ABS($I{VS})<=0.03)'], fill=GREEN_FILL))
vf.row_dimensions[VS].height = 18
nr = VS + 2
for i, (txt, kind) in enumerate([
    ("Hinweise", "h"),
    ("• Jede Zeile vergleicht einen eigenen Auslegungspunkt (Spalten A–E) mit dem Hersteller-/Referenzwert (Spalte F).", ""),
    ("• Hersteller-q (Spalte F) aus dem Datenblatt eintragen. Vorbelegt: 12 reale Referenzpunkte des Rehau 17×2-Systems", ""),
    ("   (Estrichüberdeckung s_ü = 45 mm, zwei Belag-/Temperaturfälle, Verlegeabstand 50–300 mm).", ""),
    ("• Andere nach EN 1264 zertifizierte Systeme (Uponor, Purmo, Buderus, Kermi …) liegen nah an diesen Werten.", ""),
    ("• Systemfaktor f oben so wählen, dass die Gesamtabweichung (Summenzeile) nahe 0 liegt. f ≈ 0,85 passt gut.", ""),
    ("• Einzelne Punkte können bei sehr engem/weitem Verlegeabstand abweichen (vereinfachte η-Tabelle); der", ""),
    ("   Gesamtabgleich (Summe) zeigt, dass f das Modell im Mittel sehr gut an die Herstellerdaten anpasst.", ""),
    ("• f bündelt die im überschlägigen Modell vereinfachten Effekte (Rohrwand-, Estrichspreizungs-Widerstand).", "i")]):
    c = vf.cell(row=nr + i, column=1, value=txt)
    if kind == "h": c.font = f(bold=True, color=NAVY, size=11)
    elif kind == "i": c.font = f(italic=True, color=GREY, size=9)
    else: c.font = f()
disp_header(vf, "Verifikation – Abgleich mit Herstellerdaten", 12, project=True)
setup_print(vf, f"A1:L{max(nr + 8, 22)}")

# =====================================================================
#  Blatt 6: METHODIK
# =====================================================================
m = wb.create_sheet("Methodik")
m.sheet_view.showGridLines = False
for col, w in (("A", 3), ("B", 46), ("C", 3), ("D", 48), ("E", 3), ("F", 48)):
    m.column_dimensions[col].width = w

def methcol(col, start, items):
    rr = start
    for text, kind in items:
        c = m.cell(row=rr, column=col, value=text)
        if kind == "h": c.font = Font(name=FONT, bold=True, color=NAVY, size=11)
        elif kind == "fb": c.font = Font(name=FONT, bold=True, color=BLACK, size=10)
        elif kind == "i": c.font = Font(name=FONT, italic=True, color=GREY, size=9)
        else: c.font = Font(name=FONT, size=10)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        rr += 1
    return rr

m["B4"].value = "Formelzeichen / Variablen"; m["B4"].font = Font(name=FONT, bold=True, color=NAVY, size=11)
m["D4"].value = "Berechnungsschritte"; m["D4"].font = Font(name=FONT, bold=True, color=NAVY, size=11)
m.merge_cells("D4:F4")
vars_ = [
    ("θV, θR — Vor-/Rücklauftemperatur [°C]", ""), ("θi — Raumtemperatur [°C]", ""),
    ("θ_u — Temperatur unter der Decke [°C]", ""), ("ΔθH — log. Übertemperatur [K]", ""),
    ("q — spez. Heizleistung nach oben [W/m²]", ""), ("α — Wärmeübergang Oberfläche [W/m²K]", ""),
    ("R_Belag — Widerstand Bodenbelag [m²K/W]", ""), ("s_ü — Estrichüberdeckung [m]", ""),
    ("λ_E — Wärmeleitfähigkeit Estrich [W/mK]", ""), ("K_H — Wärmedurchgang nach oben [W/m²K]", ""),
    ("η_VA — Verlegeabstandsfaktor [-]", ""), ("f — Systemfaktor (EN-1264-Kalibrierung) [-]", ""),
    ("R_u — Widerstand nach unten [m²K/W]", ""), ("q_u — Wärmestrom nach unten [W/m²]", ""),
    ("θF, θF,max — Oberflächentemp. / Grenze [°C]", ""), ("di, A_i — Rohr-Innen-Ø / -Querschnitt", ""),
    ("ρ, c_p, ν — Dichte / Wärmekap. / Viskosität", ""), ("ṁ, V̇ — Massen-/Volumenstrom [kg/h]/[l/h]", ""),
    ("v, Re, λ — Geschw. / Reynolds / Reibung", ""), ("L_Kreis — Rohrlänge je Kreis [m]", ""),
    ("Δp — Druckverlust [Pa]", ""),
]
steps_l = [
    ("1)  ΔθH = (θV−θR) / ln((θV−θi)/(θR−θi))", "fb"),
    ("Mittlere treibende Temperaturdifferenz Heizwasser ↔ Raum (log. Mittel).", "i"), ("", ""),
    ("2)  K_H = 1 / (1/α + R_Belag + s_ü/λ_E)", "fb"),
    ("Wärmedurchgang des Aufbaus nach oben.", "i"), ("", ""),
    ("3)  q = f · η_VA · K_H · ΔθH", "fb"),
    ("Spez. Heizleistung nach oben; f kalibriert an Herstellerdaten.", "i"), ("", ""),
    ("4)  θF = θi + q/α", "fb"),
    ("Bodenoberflächentemperatur; Prüfung gegen θF,max.", "i"), ("", ""),
    ("5)  Leistung oben = q · Fläche ; Deckung % = Leistung / Heizlast", "fb"),
    ("Prüft, ob die FBH die Heizlast deckt (≥ 100 %).", "i"),
]
steps_r = [
    ("6)  q_u = ((θV+θR)/2 − θ_u)/R_u ; Verlust = q_u · Fläche", "fb"),
    ("Wärmeabgabe nach unten durch die Dämmung.", "i"), ("", ""),
    ("7)  maßgebl. Leistung = MIN(Leistung ; Heizlast) + Verlust", "fb"),
    ("Tatsächlich transportierte Wärme (Basis für Massenstrom).", "i"), ("", ""),
    ("8)  L_Kreis = (Fläche/Verlegeabstand + 2·Zuleitung·Kreise)/Kreise", "fb"),
    ("Rohrlänge je Kreis; Prüfung gegen max. Kreislänge.", "i"), ("", ""),
    ("9)  ṁ = (maßgebl. Leistung/Kreise)/(c_p·(θV−θR))", "fb"),
    ("v = V̇/A_i ; Re = v·di/ν ; Re < 2300 = laminar.", "i"), ("", ""),
    ("10) Δp = λ·(L/di)·(ρ/2)·v²·(1+Zuschlag) + Verteiler", "fb"),
    ("Druckverlust je Kreis; Kontrolle gegen Warnschwelle.", "i"), ("", ""),
    ("Vereinfachungen", "h"),
    ("• Gilt nur für Bauart A nach EN 1264 (Heizrohre im Estrich).", ""),
    ("• K_H vereinfacht + η-Faktor statt EN-1264; Rohrwand nicht enthalten (→ f).", ""),
    ("• Wärmeabgabe nach unten global; pauschale Δp-Zuschläge.", ""),
]
end_b = methcol(2, 5, [(fz(t), k) for t, k in vars_])
methcol(4, 6, steps_l)
methcol(6, 6, steps_r)
disp_header(m, "Methodik & Annahmen", 6, project=False)
setup_print(m, f"A1:F{end_b}")

# =====================================================================
#  Blatt 7: KONSTANTEN
# =====================================================================
kt = wb.create_sheet("Konstanten")
kt.sheet_view.showGridLines = False
for col, w in (("A", 22), ("B", 12), ("C", 9), ("D", 17), ("E", 9), ("F", 12),
               ("G", 3), ("H", 25), ("I", 12), ("J", 2)):
    kt.column_dimensions[col].width = w
kt.cell(row=3, column=1, value="Nachschlage-Tabellen (editierbar). Leerzeilen = Platz für weitere Einträge.").font = f(italic=True, color=GREY, size=9)
# Oben nebeneinander: Verlegeabstand-Faktor (links) | Zonen (Mitte) | Leerspalte | R-Werte (rechts, höher)
two_col_table(kt, 4, 1, "Verlegeabstand-Faktor", "Verlegeabstand [mm]", "Faktor η [-]",
    [(50, 1.05), (75, 1.02), (100, 1.00), (125, 0.97), (150, 0.93), (200, 0.85), (250, 0.78), (300, 0.72)], '0.000')
# Zonen 3-spaltig (Zone | Kürzel | θF,max) – Kürzel wird in der Auslegung verwendet
kt.cell(row=4, column=4, value="Zonen / max. Oberflächentemp.").font = f(bold=True, color=NAVY)
kt.merge_cells("D4:F4")
mini_header(kt, 5, 4, "Zone"); mini_header(kt, 5, 5, "Kürzel"); mini_header(kt, 5, 6, "θF,max [°C]")
zdata = [("Aufenthaltszone", "AZ", 29), ("Randzone", "RZ", 35), ("Badezimmer", "BAD", 33)]
for i in range(8):   # 3 + 5 leer (gleiche Höhe wie Verlegeabstand-Tabelle)
    r = 6 + i
    for col in (4, 5, 6):
        cell = kt.cell(row=r, column=col); cell.font = f(color=BLUE); cell.fill = INPUT_FILL; cell.border = BORDER
        cell.alignment = Alignment(horizontal="left" if col <= 5 else "center")
        if col == 6: cell.number_format = '0" °C"'
        if i < len(zdata): cell.value = zdata[i][col - 4]
# Spalte G bleibt als Leerspalte frei (Trennung Zonen ↔ R-Werte)
# R-Werte Bodenbeläge (rechts ab Spalte H, mehr Auswahl). R-Wert in Spalte I = Dropdown-Quelle der Auslegung.
two_col_table(kt, 4, 8, "R-Werte Bodenbeläge", "Bodenbelag", "R [m²·K/W]", BELAEGE, '0.000', n_empty=3)
# Darunter volle Breite: Rohrbibliothek
kt.cell(row=16, column=1, value="Rohrbibliothek (di = da − 2·s)").font = f(bold=True, color=NAVY)
for j, h in enumerate(["Rohrsystem", "da [mm]", "s [mm]", "di [mm]", "k [mm]"], start=1):
    mini_header(kt, 17, j, h)
pipes = [("PE-Xa 17x2", 17, 2.0, 0.007), ("PE-Xa 16x2", 16, 2.0, 0.007),
         ("PE-Xa 20x2", 20, 2.0, 0.007), ("PE-RT 14x2", 14, 2.0, 0.007)]
colmap = {1: 0, 2: 1, 3: 2, 5: 3}
for i in range(8):
    r = 18 + i
    for col in range(1, 6):
        cell = kt.cell(row=r, column=col); cell.border = BORDER
        if col == 4:
            cell.value = f'=IF(B{r}="","",B{r}-2*C{r})'; cell.font = f(color=BLACK); cell.number_format = '0.0'; cell.alignment = CEN
        else:
            cell.font = f(color=BLUE); cell.fill = INPUT_FILL
            cell.alignment = Alignment(horizontal="left" if col == 1 else "center")
            if col in (2, 3): cell.number_format = '0.0'
            if col == 5: cell.number_format = '0.000'
            if i < len(pipes): cell.value = pipes[i][colmap[col]]
disp_header(kt, "Konstanten / Bibliotheken", 9, project=False)
setup_print(kt, "A1:I27")

# =====================================================================
#  Blatt 8: CHANGELOG
# =====================================================================
cl = wb.create_sheet("Changelog")
cl.sheet_view.showGridLines = False
for col, w in (("A", 12), ("B", 14), ("C", 115)):
    cl.column_dimensions[col].width = w
for j, h in enumerate(["Version", "Datum", "Änderungen"], start=1):
    c = cl.cell(row=5, column=j, value=h)
    c.font = f(bold=True, color=WHITE); c.fill = HDR_FILL
    c.alignment = Alignment(horizontal="left", vertical="center"); c.border = BORDER
r = 6
for ver, datum, changes in reversed(CHANGELOG):
    start = r
    for ch in changes:
        cl.cell(row=r, column=3, value="• " + ch).font = f()
        cl.cell(row=r, column=3).alignment = Alignment(wrap_text=True, vertical="top")
        for col in (1, 2, 3): cl.cell(row=r, column=col).border = BORDER
        r += 1
    a = cl.cell(row=start, column=1, value=ver); a.font = f(bold=True); a.alignment = Alignment(vertical="top")
    b = cl.cell(row=start, column=2, value=datum); b.font = f(); b.alignment = Alignment(vertical="top")
    if r - start > 1:
        cl.merge_cells(start_row=start, start_column=1, end_row=r - 1, end_column=1)
        cl.merge_cells(start_row=start, start_column=2, end_row=r - 1, end_column=2)
disp_header(cl, "Changelog – FBH-Auslegung", 3, project=False)
cl.cell(row=3, column=3, value=VERTXT).font = f(italic=True, color=GREY, size=9)
cl.cell(row=3, column=3).alignment = Alignment(horizontal="right")
setup_print(cl, f"A1:C{r-1}")

# =====================================================================
#  Blatt 0: DECKBLATT  (Übersichtsdiagramm aus Zellen – in Excel editierbar)
# =====================================================================
db = wb.create_sheet("Deckblatt", 0)
db.sheet_view.showGridLines = False
for col, w in (("A", 14), ("B", 14), ("C", 14), ("D", 5), ("E", 16), ("F", 16),
               ("G", 16), ("H", 5), ("I", 14), ("J", 14), ("K", 14)):
    db.column_dimensions[col].width = w
RED_S = PatternFill("solid", fgColor="E2001A")
GREEN_S = PatternFill("solid", fgColor="2E7D32")
BODY_S = PatternFill("solid", fgColor="F7F7F7")
AMB_H = PatternFill("solid", fgColor="FFE699")
AMB_B = PatternFill("solid", fgColor="FBF3D9")
AMBER = "B88600"

def box(r1, c1, r2, c2, value=None, fill=None, font=None, align="left", valign="top", border=True):
    db.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cc = db.cell(row=r, column=c)
            if fill: cc.fill = fill
            if border: cc.border = BORDER
    tl = db.cell(row=r1, column=c1)
    if value is not None: tl.value = value
    if font: tl.font = font
    tl.alignment = Alignment(horizontal=align, vertical=valign, wrap_text=True)

db.row_dimensions[1].height = 28
db.merge_cells("A1:H1")
db["A1"].value = "Fußbodenheizung – Auslegung"
db["A1"].font = Font(name=FONT, bold=True, color=NAVY, size=18); db["A1"].alignment = Alignment(vertical="center")
db.merge_cells("A2:H2"); db["A2"].value = "Eingaben · Verarbeitung · Ergebnisse"; db["A2"].font = f(color=GREY, size=12)
db.merge_cells("A3:K3")
db["A3"].value = f'="Projekt-Nr.: "&{gPNr}&"          Projekt: "&{gPName}&"          Bearbeiter: "&{gBearb}'
db["A3"].font = f(bold=True, color=NAVY, size=11)
add_logo_corner(db, 11, height=34)

# Spalten-Köpfe (fett)
box(5, 1, 5, 3, "EINGABEN", fill=HDR_FILL, font=f(bold=True, color=WHITE, size=13), align="center", valign="center")
box(5, 5, 5, 7, "BERECHNUNG", fill=RED_S, font=f(bold=True, color=WHITE, size=13), align="center", valign="center")
box(5, 9, 5, 11, "ERGEBNISSE", fill=GREEN_S, font=f(bold=True, color=WHITE, size=13), align="center", valign="center")

# Formelzeichen mit echten Tiefstellungen (für die Berechnungs-Spalte)
def rsub(*segs):
    sf = InlineFont(rFont=FONT, vertAlign="subscript", b=True, sz=10, color=NAVY)
    return CellRichText([TextBlock(sf, t) if sub else t for t, sub in segs])

# Eine Information = eine Zelle (keine zusammenhängenden Mehrzeilen-Blöcke, keine Fehl-Umbrüche)
def db_item(r, c1, c2, text, kind):
    if kind == "h":        # Zwischenüberschrift
        box(r, c1, r, c2, text, fill=SUB_FILL, font=f(bold=True, color=NAVY, size=10.5), align="left", valign="center")
    elif kind == "fmla":   # Formel (Rich-Text mit Tiefstellung) – klar sichtbar in der Berechnungs-Spalte
        box(r, c1, r, c2, text, fill=None, font=f(bold=True, color=NAVY, size=10.5), align="left", valign="center")
    elif kind == "b":      # Aufzählungspunkt
        box(r, c1, r, c2, "•  " + text, fill=BODY_S, font=f(color=BLACK, size=10), align="left", valign="center")
    else:                  # Leerzeile (Auffüllung für gleiche Spaltenhöhe)
        box(r, c1, r, c2, "", fill=BODY_S, align="left", valign="center")

def db_col(c1, c2, items, start=6):
    for i, (text, kind) in enumerate(items):
        db_item(start + i, c1, c2, text, kind)

eingaben = [
    ("Globale Vorgaben", "h"),
    ("Vorlauf- und Rücklauftemperatur", "b"),
    ("Rohrsystem und Stoffwerte", "b"),
    ("Wärmeübergang und Estrichaufbau", "b"),
    ("Dämmung nach unten", "b"),
    ("Grenzwerte (Länge, Druck, Strömung)", "b"),
    ("Je Raum", "h"),
    ("Raumfläche und aktivierbare Fläche", "b"),
    ("Heizlast und Raumtemperatur", "b"),
    ("Bodenbelag-Widerstand, Verlegeabstand", "b"),
    ("Anzahl Heizkreise, Zuleitung, Zone", "b"),
    ("Heizkreisverteiler", "b"),
    ("Kalibrierung", "h"),
    ("Systemfaktor (DIN EN 1264 / Hersteller)", "b"),
]
berechnung = [
    ("Heizleistung nach oben", "h"),
    (rsub(("q = f · η · K", False), ("H", True), (" · Δθ", False), ("H", True)), "fmla"),
    (rsub(("Δθ", False), ("H", True), (" = (θ", False), ("V", True), (" − θ", False), ("R", True),
          (") / ln[(θ", False), ("V", True), (" − θ", False), ("i", True), (") / (θ", False),
          ("R", True), (" − θ", False), ("i", True), (")]", False)), "fmla"),
    (rsub(("K", False), ("H", True), (" = 1 / (1/α + R", False), ("Belag", True), (" + s", False),
          ("ü", True), (" / λ", False), ("E", True), (")", False)), "fmla"),
    ("Oberflächentemperatur", "h"),
    (rsub(("θ", False), ("F", True), (" = θ", False), ("i", True), (" + q / α", False)), "fmla"),
    ("Hydraulik / Druckverlust", "h"),
    ("Maßgebliche Leistung → Massenstrom", "b"),
    (rsub(("Δp = λ · (L / d", False), ("i", True), (") · (ρ / 2) · v²", False)), "fmla"),
    ("Strömung: Reynolds-Zahl, Rohrreibung", "b"),
    ("Rohrlängen je Heizkreis", "b"),
    ("Prüfungen (Ampelbewertung)", "h"),
    ("Deckung, Temperatur, Druckverlust", "b"),
    ("", ""),
]
ergebnisse = [
    ("Je Raum", "h"),
    ("Spezifische Heizleistung", "b"),
    ("Oberflächentemperatur", "b"),
    ("Deckungsgrad (Ampelbewertung)", "b"),
    ("Rohrlänge je Heizkreis (Ampel)", "b"),
    ("Massenstrom und Volumenstrom", "b"),
    ("Strömungsgeschwindigkeit", "b"),
    ("Druckverlust je Heizkreis (Ampel)", "b"),
    ("Je Heizkreisverteiler", "h"),
    ("Leistung und Volumenstrom", "b"),
    ("Maximaler Druckverlust", "b"),
    ("Gesamt", "h"),
    ("Kontrolle und Warnübersicht", "b"),
    ("", ""),
]
db_col(1, 3, eingaben)
db_col(5, 7, berechnung)
db_col(9, 11, ergebnisse)
NROW = max(len(eingaben), len(berechnung), len(ergebnisse))
last = 6 + NROW - 1
# Pfeile zwischen den Spalten
box(6, 4, last, 4, "→", font=f(bold=True, color="E2001A", size=22), align="center", valign="center", border=False)
box(6, 8, last, 8, "→", font=f(bold=True, color="E2001A", size=22), align="center", valign="center", border=False)
db.row_dimensions[5].height = 20
for r in range(6, last + 1): db.row_dimensions[r].height = 18

# Fußzeile: Formelzeichen-Legende + Verweis auf Methodik
box(last + 2, 1, last + 2, 11,
    "Formelzeichen:  θV / θR  Vorlauf- / Rücklauftemperatur · θi  Raumtemperatur · θF  Oberflächentemperatur · "
    "α  Wärmeübergangskoeffizient · η  Verlegeabstand-Faktor · f  Systemfaktor · L  Rohrlänge · "
    "di  Innendurchmesser · λ  Rohrreibungszahl · ρ  Dichte · v  Strömungsgeschwindigkeit",
    fill=None, font=f(italic=True, color=GREY, size=9), align="left", valign="center", border=False)
box(last + 3, 1, last + 3, 11,
    "Überschlägige Auslegung – Vereinfachungen und Annahmen siehe Blatt „Methodik“.",
    fill=None, font=f(italic=True, color=GREY, size=9), align="left", valign="center", border=False)
db.page_setup.orientation = "landscape"
db.page_setup.paperSize = 9
db.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
db.page_setup.fitToWidth = 1; db.page_setup.fitToHeight = 1
db.print_area = f"A1:K{last + 3}"
db.page_margins.left = db.page_margins.right = 0.3
db.page_margins.top = db.page_margins.bottom = 0.3

# =====================================================================
#  Blatt 1: ANLEITUNG  (zweites Blatt)
# =====================================================================
an = wb.create_sheet("Anleitung", 1)
an.sheet_view.showGridLines = False
for col, w in (("A", 2), ("B", 60), ("C", 3), ("D", 60)):
    an.column_dimensions[col].width = w

def anl(col, start, items):
    rr = start
    for text, kind in items:
        c = an.cell(row=rr, column=col, value=text)
        if kind == "h": c.font = Font(name=FONT, bold=True, color=NAVY, size=12)
        elif kind == "sh": c.font = Font(name=FONT, bold=True, color=BLACK, size=10.5)
        elif kind == "i": c.font = Font(name=FONT, italic=True, color=GREY, size=9)
        else: c.font = Font(name=FONT, size=10)
        c.alignment = Alignment(vertical="top")
        rr += 1
    return rr

left = [
    ("Aufbau der Mappe", "h"),
    ("• Deckblatt – Übersicht (Input / Verarbeitung / Output)", ""),
    ("• Grundeinstellungen – globale Werte", ""),
    ("• Auslegung – Raumliste mit allen Berechnungen", ""),
    ("• Kontrolle – Summen & Warnungen", ""),
    ("• HKV – Auswertung je Heizkreisverteiler", ""),
    ("• Verifikation – Abgleich mit Herstellerdaten", ""),
    ("• Anleitung – diese Seite", ""),
    ("• Methodik – Formeln & Annahmen", ""),
    ("• Konstanten – Nachschlagetabellen", ""),
    ("• Changelog – Versionen", ""),
    ("", ""),
    ("Farblegende", "h"),
    ("• Blau (gelb hinterlegt) = Eingabe", ""),
    ("• Schwarz = berechnet (nicht ändern)", ""),
    ("• Grün = OK,  Rot = Warnung / Grenzwert überschritten", ""),
    ("• Nur die blauen Eingabezellen bearbeiten.", ""),
    ("", ""),
    ("Ampel-Legende (warum ist eine Zelle rot?)", "h"),
    ("• Deckung < 100 % → Heizlast nicht gedeckt", ""),
    ("• Oberflächentemperatur > Zonen-Grenztemperatur", ""),
    ("• Rohrlänge je Kreis > max. Kreislänge", ""),
    ("• Druckverlust je Kreis > Warnschwelle", ""),
    ("• Geschwindigkeit außerhalb der Grenzen", ""),
    ("• Volumenstrom je Kreis > Grenzwert", ""),
    ("• aktivierbare Fläche > Raumfläche", ""),
    ("Grenzwerte stehen in den Grundeinstellungen,", "i"),
    ("Zonen-Grenztemperaturen in den Konstanten.", "i"),
]
right = [
    ("Schritt für Schritt", "h"),
    ("1. Grundeinstellungen ausfüllen: Temperaturen, Rohr-", ""),
    ("    system, Stoffwerte, Grenzwerte.", ""),
    ("2. Konstanten prüfen / erweitern (Rohre, R-Werte, Zonen).", ""),
    ("3. Verifikation: Systemfaktor f kalibrieren – VOR der", ""),
    ("    Raumauslegung (er wirkt auf alle Räume).", ""),
    ("4. Bauteilliste (Schedule) aus Revit nach Excel exportieren", ""),
    ("    – Spalten A–F (HKV, Raum-Nr., Bezeichnung, Raum-", ""),
    ("    fläche, Raumtemperatur, Heizlast).", ""),
    ("5. Export in die Auslegung übernehmen (A–F).", ""),
    ("6. Spalte G (spez. Heizlast) frei lassen – wird berechnet.", ""),
    ("7. Spalten H–M je Raum ergänzen (aktiv. Fläche, R-Wert,", ""),
    ("    Verlegeabstand, Kreise, Zuleitung, Zone).", ""),
    ("8. Ergebnisse prüfen: Ampel in Auslegung + 'Kontrolle'.", ""),
    ("", ""),
    ("Spalten der Auslegung", "h"),
    ("Aus Revit (A–F): HKV, Raum-Nr., Bezeichnung,", ""),
    ("   Raumfläche, Raumtemperatur, Heizlast.", ""),
    ("Berechnet (G): spez. Heizlast (nicht ändern).", ""),
    ("Im Tool (H–M): aktiv. Fläche, R-Wert, Verlege-", ""),
    ("   abstand, Kreise, Zuleitung, Zone.", ""),
    ("", ""),
    ("Mehrere Verteiler / Geschosse auslegen", "h"),
    ("• Verteiler je Raum in Spalte A (HKV) eintragen", ""),
    ("   (z. B. HKV EG, HKV OG, HKV 1.OG …).", ""),
    ("• Das Blatt 'HKV' wertet jeden Verteiler automatisch", ""),
    ("   aus (Kreise, Leistung, Volumenstrom, max. Δp).", ""),
    ("• Beliebig viele Verteiler in EINER Liste – einfach", ""),
    ("   weitere Räume ergänzen.", ""),
]
end_l = anl(2, 3, left)
end_r = anl(4, 3, right)
disp_header(an, "Anleitung / Bedienung", 4, project=False)
setup_print(an, f"A1:D{max(end_l, end_r)}")

# Blattreihenfolge: Anleitung nach hinten, direkt vor die Methodik
SHEET_ORDER = ["Deckblatt", "Grundeinstellungen", "Auslegung", "Kontrolle", "HKV",
               "Verifikation", "Anleitung", "Methodik", "Konstanten", "Changelog"]
wb._sheets.sort(key=lambda ws: SHEET_ORDER.index(ws.title))
wb.active = 0

# =====================================================================
# Blattschutz: nur Formelzellen sperren (Eingabezellen bleiben editierbar).
# Schutz ohne Passwort → kann bei Bedarf über 'Überprüfen ▸ Blattschutz aufheben' gelöst werden.
def protect_formulas(ws):
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            v = cell.value
            is_formula = isinstance(v, ArrayFormula) or (isinstance(v, str) and v.startswith("="))
            cell.protection = Protection(locked=is_formula)
    ws.protection.sheet = True
    # Zeilenhöhe/Spaltenbreite trotz Blattschutz änderbar lassen (z. B. HKV-Raumliste umbrechen)
    ws.protection.formatRows = False
    ws.protection.formatColumns = False
for ws in wb.worksheets:
    protect_formulas(ws)

# =====================================================================
wb.properties.version = VERSION
wb.properties.creator = AUTHOR
wb.properties.lastModifiedBy = AUTHOR
wb.properties.keywords = f"FBH-Auslegung v{VERSION} ({AUTHOR})"
try:
    wb.calculation.fullCalcOnLoad = True
except Exception:
    pass
OUT = r"C:\Users\derDi\FBH_Auslegung_Excel\FBH_Auslegung.xlsx"
wb.save(OUT)
print("gespeichert:", OUT, "| Version", VERSION, AUTHOR)
