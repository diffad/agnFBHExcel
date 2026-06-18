# -*- coding: utf-8 -*-
"""Minimalbeispiel für den agn-Graustufen-Hausstil (siehe agn_xlsx_style.py).
Baut eine kleine Mappe: Deckblatt + Datenblatt (Eingabe weiß / berechnet grau, Zebra,
Bewertung als Schriftfarbe) + Changelog. Als Kopiervorlage gedacht."""
import os, sys
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from openpyxl import Workbook
import agn_xlsx_style as S

LOGO = r"C:\Users\derDi\FBH_Auslegung_Excel\agn-logo.png"   # optional; fehlt → einfach kein Logo
R0, R1 = 8, 40   # Datenbereich

wb = Workbook()

# ---- Datenblatt ----
ws = wb.active; ws.title = "Daten"; ws.sheet_view.showGridLines = False
for col, w in (("A", 18), ("B", 14), ("C", 14), ("D", 14)):
    ws.column_dimensions[col].width = w
S.letterhead(ws, "Beispiel-Tabelle", last_col=4, logo_path=LOGO)
cols = [("Bezeichnung", None), ("Eingabe", "[-]"), ("berechnet", "[-]"), ("Grenzwert", "[-]")]
for j, (name, unit) in enumerate(cols, start=1):
    S.header_cell(ws, 6, j, name)
    S.header_cell(ws, 7, j, unit or "", underline=True)
demo = [("Raum 1", 80), ("Raum 2", 120), ("Raum 3", 95), ("Raum 4", 150)]
for i, r in enumerate(range(R0, R1 + 1)):
    a = ws.cell(row=r, column=1, value=demo[i][0] if i < len(demo) else None); S.style_input(a, "left")
    b = ws.cell(row=r, column=2, value=demo[i][1] if i < len(demo) else None); S.style_input(b)
    c = ws.cell(row=r, column=3, value=f'=IF($B{r}="","",$B{r}*1.1)'); S.style_calc(c)
    d = ws.cell(row=r, column=4, value=f'=IF($B{r}="","",110)'); S.style_calc(d)
ws.freeze_panes = "A8"
ws.auto_filter.ref = f"A7:D{R1}"
# Bewertung Spalte C (grün wenn <= Grenzwert D, sonst rot) + Zebra kombiniert
S.cf_eval(ws, "C", R0, R1,
          ok_formula=f'AND($C{R0}<>"",$C{R0}<=$D{R0})',
          bad_formula=f'AND($C{R0}<>"",$C{R0}>$D{R0})', keycol="A", calc=True)
# Zebra für die übrigen Spalten (A weiß, B weiß, D grau)
S.zebra(ws, f"A{R0}:A{R1}", "A", R0, calc=False)
S.zebra(ws, f"B{R0}:B{R1}", "A", R0, calc=False)
S.zebra(ws, f"D{R0}:D{R1}", "A", R0, calc=True)
S.setup_print(ws, f"A1:D{R1}", titles="1:7", footer_left="Beispiel")

# ---- Changelog ----
cl = wb.create_sheet("Changelog"); cl.sheet_view.showGridLines = False
for col, w in (("A", 12), ("B", 14), ("C", 80)):
    cl.column_dimensions[col].width = w
S.letterhead(cl, "Changelog", last_col=3, logo_path=LOGO)
for j, h in enumerate(["Version", "Datum", "Änderungen"], start=1):
    S.header_cell(cl, 6, j, h, underline=True)
S.style_calc(cl.cell(row=7, column=1, value="0.1")); S.style_calc(cl.cell(row=7, column=2, value="2026-06-17"))
S.style_calc(cl.cell(row=7, column=3, value="Ersterstellung im agn-Hausstil."), "left")
S.setup_print(cl, "A1:C7")

for sheet in wb.worksheets:
    S.lock_formulas(sheet, allow_filter=(sheet is ws))

OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "agn_xlsx_example.xlsx")
wb.save(OUT)
print("gespeichert:", OUT)
