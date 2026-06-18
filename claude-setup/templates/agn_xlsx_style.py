# -*- coding: utf-8 -*-
"""
agn_xlsx_style — wiederverwendbarer Graustufen-Hausstil für Excel-Arbeitsmappen (openpyxl).

Designprinzipien:
- Graustufen; Schrift schwarz. Farbe NUR im Logo und in der Bewertung (grün = ok / rot = Grenzwert),
  und zwar als SCHRIFTFARBE, nicht als Zellhintergrund.
- Tabellenkopf einheitlich kräftiges Grau (HEADER), darunter eine schwarze Linie.
- BERECHNETE Zellen leicht grau (CALC), EINGABE-/editierbare Zellen weiß (INPUT).
- Haarlinien statt Vollgitter. Zebra-Streifen (jede zweite sichtbare Zeile) – filterfest via SUBTOTAL.
- Cross-App: läuft in Excel UND LibreOffice (kein Spill/UNIQUE; TEXTJOIN mit _xlfn.-Präfix).
- Kein VBA/Makro (Sicherheitsvorgabe). Formelzellen werden gesperrt, Eingaben bleiben editierbar.

Empfohlener Mappen-Aufbau (Blattreihenfolge):
  Deckblatt · Anleitung · Grundeinstellungen · <Auslegung/Daten> · Kontrolle · <Auswertungen> ·
  Verifikation · Methodik · Konstanten · Changelog

Minimalbeispiel:
    from openpyxl import Workbook
    import agn_xlsx_style as S
    wb = Workbook(); ws = wb.active; ws.title = "Daten"; ws.sheet_view.showGridLines = False
    S.letterhead(ws, "Titel der Mappe", last_col=4, logo_path=r"C:\\pfad\\agn-logo.png")
    # ... Kopf + Zellen mit S.HEADER_FILL / S.style_input / S.style_calc, dann:
    S.zebra(ws, "A8:D200", keycol="A", top=8, calc=True)
    S.setup_print(ws, "A1:D200"); S.lock_formulas(ws, allow_filter=True)
    wb.save("Mappe.xlsx")
"""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.cell.cell import MergedCell
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.utils import get_column_letter

FONT = "Arial"

# ---- Palette (Graustufen) ----
INK        = "000000"   # Schrift (immer schwarz)
GREY_NOTE  = "808080"   # Hinweise/Einheiten-Sekundärtext (sparsam)
HEADER     = "D2D2D2"   # Tabellenkopf / Abschnittsbalken (kräftiges Grau)
SUBHEAD    = "E6E6E6"
CALC       = "ECECEC"   # berechnete Zellen (leichtes Grau)
INPUT      = "FFFFFF"   # Eingabe-/editierbare Zellen (weiß)
HAIRLINE   = "C9C9C9"
ZEBRA_INPUT_C = "F3F3F3"  # Streifen auf weißen Eingabespalten
ZEBRA_CALC_C  = "E0E0E0"  # Streifen auf grauen Rechenspalten
EVAL_OK    = "1E8E1E"   # Bewertung grün (im Rahmen)
EVAL_WARN  = "D32F2F"   # Bewertung rot (Grenzwert überschritten)

HEADER_FILL = PatternFill("solid", fgColor=HEADER)
SUBHEAD_FILL = PatternFill("solid", fgColor=SUBHEAD)
CALC_FILL   = PatternFill("solid", fgColor=CALC)
INPUT_FILL  = PatternFill("solid", fgColor=INPUT)
ZEBRA_INPUT = PatternFill(start_color="FF" + ZEBRA_INPUT_C, end_color="FF" + ZEBRA_INPUT_C, fill_type="solid")
ZEBRA_CALC  = PatternFill(start_color="FF" + ZEBRA_CALC_C, end_color="FF" + ZEBRA_CALC_C, fill_type="solid")

_hair = Side(style="thin", color=HAIRLINE)
HAIR_BORDER = Border(left=_hair, right=_hair, top=_hair, bottom=_hair)
HEAD_UNDERLINE = Border(left=_hair, right=_hair, top=_hair, bottom=Side(style="medium", color=INK))
CEN, LEFT = Alignment(horizontal="center"), Alignment(horizontal="left")


def font(bold=False, color=INK, size=10, italic=False):
    return Font(name=FONT, bold=bold, color=color, size=size, italic=italic)


def letterhead(ws, title, last_col, logo_path=None, project_formula=None, logo_height=36):
    """Briefkopf: schwarzer Titel links, Logo rechts, dünne schwarze Trennlinie unter dem Kopf.
    project_formula z. B. '="Projekt-Nr.  "&Grundeinstellungen!$C$2'."""
    ws.row_dimensions[1].height = 32
    t = ws.cell(row=1, column=1, value=title)
    t.font = Font(name=FONT, bold=True, color=INK, size=16)
    t.alignment = Alignment(horizontal="left", vertical="center")
    if logo_path:
        _logo(ws, last_col, logo_path, logo_height)
    if project_formula:
        m = ws.cell(row=2, column=1, value=project_formula); m.font = font(size=11); m.alignment = LEFT
    line = Border(top=Side(style="thin", color=INK))
    for c in range(1, last_col + 1):
        ws.cell(row=3, column=c).border = line
    ws.row_dimensions[3].height = 8


def _logo(ws, last_col, path, height):
    import os
    if not os.path.exists(path):
        return
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
    from openpyxl.drawing.xdr import XDRPositiveSize2D
    EMU = 9525
    def col_px(i):
        dim = ws.column_dimensions[get_column_letter(i)]
        w = dim.width if dim.width else 8.43
        return 0 if dim.hidden else int(round(w * 7)) + 5
    img = XLImage(path); ratio = img.width / img.height
    img.height = height; img.width = int(height * ratio)
    right = sum(col_px(i) for i in range(1, last_col + 1)); left = max(0, right - img.width - 4)
    acc, c = 0, 1
    while c <= last_col:
        cw = col_px(c)
        if cw and acc + cw > left:
            break
        acc += cw; c += 1
    img.anchor = OneCellAnchor(_from=AnchorMarker(col=c - 1, colOff=(left - acc) * EMU, row=0, rowOff=EMU),
                               ext=XDRPositiveSize2D(cx=img.width * EMU, cy=img.height * EMU))
    ws.add_image(img)


def section_bar(ws, row, col_start, col_end, text):
    """Grauer Abschnittsbalken, schwarze fette Schrift (KEINE weiße Schrift)."""
    for c in range(col_start, col_end + 1):
        ws.cell(row=row, column=c).fill = HEADER_FILL
    cell = ws.cell(row=row, column=col_start, value=text)
    cell.font = font(bold=True, color=INK)


def header_cell(ws, row, col, text, underline=False):
    """Eine Tabellenkopf-Zelle: einheitlich grau, schwarze fette Schrift, zentriert."""
    c = ws.cell(row=row, column=col, value=text)
    c.fill = HEADER_FILL; c.font = font(bold=True, color=INK)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = HEAD_UNDERLINE if underline else HAIR_BORDER
    return c


def style_input(cell, align="center"):
    """Editierbare Zelle: weiß, schwarz, Haarlinie."""
    cell.font = font(color=INK); cell.fill = INPUT_FILL; cell.border = HAIR_BORDER
    cell.alignment = Alignment(horizontal=align)
    return cell


def style_calc(cell, align="center"):
    """Berechnete Zelle: leicht grau, schwarz, Haarlinie."""
    cell.font = font(color=INK); cell.fill = CALC_FILL; cell.border = HAIR_BORDER
    cell.alignment = Alignment(horizontal=align)
    return cell


def _odd(keycol, top):
    return f'MOD(SUBTOTAL(103,${keycol}${top}:${keycol}{top}),2)=0'


def zebra(ws, rng, keycol, top, calc=True):
    """Filterfestes Zebra (jede zweite SICHTBARE Zeile). calc=True → graue Streifen, sonst helle."""
    fill = ZEBRA_CALC if calc else ZEBRA_INPUT
    form = f'AND(${keycol}{top}<>"",{_odd(keycol, top)})'
    ws.conditional_formatting.add(rng, FormulaRule(formula=[form], fill=fill))


def cf_eval(ws, col, top, bottom, ok_formula, bad_formula, keycol="A", calc=True):
    """Bewertung als Schriftfarbe (grün ok / rot Warnung) UND filterfestes Zebra in EINER Regel.
    Wichtig: Excel/LibreOffice wenden je Zelle nur die oberste passende CF-Regel an – deshalb müssen
    Schriftfarbe und Zebra-Füllung gemeinsam gesetzt werden.
    ok_/bad_formula sind Bedingungen relativ zur Zeile `top`, z. B. f'AND($T{top}<>"",$T{top}>=$F{top})'."""
    rng = f"{col}{top}:{col}{bottom}"
    zfill = ZEBRA_CALC if calc else ZEBRA_INPUT
    odd = _odd(keycol, top)
    add = ws.conditional_formatting.add
    add(rng, FormulaRule(formula=[f'AND({bad_formula},{odd})'], font=Font(color=EVAL_WARN), fill=zfill))
    add(rng, FormulaRule(formula=[f'AND({ok_formula},{odd})'], font=Font(color=EVAL_OK), fill=zfill))
    add(rng, FormulaRule(formula=[bad_formula], font=Font(color=EVAL_WARN)))
    add(rng, FormulaRule(formula=[ok_formula], font=Font(color=EVAL_OK)))


def setup_print(ws, area, titles=None, orientation="landscape", paper=9, footer_left=""):
    """Druck auf eine Seite breit skaliert (A4=9, A3=8). titles z. B. '1:6' für Wiederholzeilen."""
    ws.page_setup.orientation = orientation
    ws.page_setup.paperSize = paper
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 0
    ws.print_area = area
    if titles:
        ws.print_title_rows = titles
    ws.page_margins.left = ws.page_margins.right = 0.3
    ws.page_margins.top = ws.page_margins.bottom = 0.4
    if footer_left:
        ws.oddFooter.left.text = footer_left
    ws.oddFooter.center.text = "Seite &P von &N"
    ws.oddFooter.right.text = "Stand &D"


def lock_formulas(ws, allow_filter=False):
    """Nur Formelzellen sperren (Eingaben bleiben editierbar); Schutz ohne Passwort.
    Zeilenhöhe/Spaltenbreite bleiben änderbar; allow_filter=True lässt Filtern trotz Schutz zu."""
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            v = cell.value
            is_formula = isinstance(v, ArrayFormula) or (isinstance(v, str) and v.startswith("="))
            cell.protection = Protection(locked=is_formula)
    ws.protection.sheet = True
    ws.protection.formatRows = False
    ws.protection.formatColumns = False
    if allow_filter:
        ws.protection.autoFilter = False
        ws.protection.sort = False


# Stolperfallen-Merkliste (Excel + LibreOffice):
#   - dxf-Füllungen brauchen FF-Alpha-Präfix (FFxxxxxx), sonst unsichtbar in Excel.
#   - Zelltexte, die mit "=" beginnen, werden als Formel interpretiert (Fehler:501) – Notizen umformulieren.
#   - TEXTJOIN/Funktionen ab Excel 2007 brauchen im XML das Präfix _xlfn. (sonst #NAME?).
#   - Bewertung NICHT als Zellhintergrund (Schwarz-Weiß-Druck) – Schriftfarbe + ✓/!-Symbol sind robuster;
#     hier bewusst Schriftfarbe gewählt (Farbdruck/Graustufen).
