# -*- coding: utf-8 -*-
"""Aufbau der Excel-Arbeitsmappe zur ueberschlaegigen Auslegung von Fussbodenheizungen."""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.cell.cell import MergedCell
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.chart import LineChart, Reference
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont

CHANGELOG = [
    ("0.1", "2026-06-16", [
        "Eigenständige Mappe: vollständige Auslegung nach dem Berechnungsansatz der DIN EN 1264-2 (Bauart A, Heizrohre im Estrich).",
        "Wärmestromdichte über das Produktverfahren q = B · a_B · a_T^mT · a_ü^mü · a_D^mD · ΔθH (kein pauschaler Systemfaktor f mehr).",
        "Faktoren a_B(R) und a_T(R) als editierbare Tabelle (an die veröffentlichten EN-1264-Leistungstabellen kalibriert: fbh24 max. 1,2 %, Rehau 17×2 −1,7 % gesamt).",
        "Exponenten m_T = 1 − Verlegeabstand/0,075, m_ü = 100·(0,045 − s_ü), m_D = 250·(D − 0,020) aus den Geometriedaten; Systemkoeffizient B = 6,7 W/m²K.",
        "Hydraulik (Massenstrom, Druckverlust nach Darcy-Weisbach), HKV-Auswertung und Kontrolle wie im vereinfachten Tool.",
        "Hinweis: a_B/a_T sind kalibriert, a_ü/a_D EN-1264-orientierte Richtwerte – bei Bedarf gegen die Norm-Tabelle (DIN EN 1264-2, Tab. A.1) prüfen.",
    ]),
    ("0.2", "2026-06-16", [
        "EN-1264-Parameter (B, a_ü, a_D, m_ü, m_D) von den Grundeinstellungen ins Blatt 'Konstanten' verschoben – alle EN-1264-Faktoren stehen jetzt zusammen.",
        "Konstanten neu geordnet: jede Tabelle durch weiße Leerzellen/-spalten optisch getrennt; passt auf eine A4-Quer-Seite.",
        "Verifikation/Hinweise sprachlich bereinigt (kein Verweis mehr auf einen Systemfaktor; Faktor-Werte können durch die Norm-Werte ersetzt werden).",
    ]),
    ("0.3", "2026-06-17", [
        "Diese EN-1264-Mappe ist jetzt die Hauptvariante; die vereinfachte Mappe wurde nach 'archiv/' verschoben.",
        "Auslegung im Graustufen-Design: Schrift durchgehend schwarz (auch Einheiten), Eingabe-/editierbare Zellen grau hinterlegt; Farbe nur noch im Logo.",
        "Bewertung als Symbol statt Farbe: neue Spalte 'Status' mit ✓ (alle Prüfungen ok) bzw. ! (Warnung) – auch im Schwarz-Weiß-Druck erkennbar; rote/grüne Zellhintergründe entfallen.",
        "Zebra-Streifen (jede zweite sichtbare Zeile) über bedingte Formatierung mit SUBTOTAL (filterfest); neue Filterfunktion (AutoFilter), Filtern trotz Blattschutz möglich.",
    ]),
    ("0.4", "2026-06-17", [
        "Graustufen-Design auf ALLE Blätter übertragen: dunkelblaue Flächen/Schrift → Grau bzw. Schwarz, gelbe Eingabefelder → helles Grau.",
        "Auslegung: Status-Spalte (✓/!) wieder entfernt; Bewertung jetzt über die Schriftfarbe – leicht grün = im Rahmen, leicht rot = Grenzwert überschritten (auch in Kontrolle und Verifikation).",
        "Tabellenkopf einheitlich in kräftigerem Grau; editierbare Zellen in hellerem Grau.",
    ]),
    ("0.5", "2026-06-17", [
        "Färbung getauscht: berechnete Zellen jetzt mit leichtem grauen Hintergrund, Eingabezellen weiß (alle Blätter).",
        "Tabellenköpfe durchgehend einheitlich grau (auch Verifikation) – keine weiße Schrift mehr, kein Unterschied zwischen Eingabe- und Rechenspalten im Kopf.",
        "Bewertungs-Schriftfarben kräftiger (grün/rot); Zebra-Streifen gelten auch für die Bewertungsspalten (rot/grüner Text).",
    ]),
]
VERSION = CHANGELOG[-1][0]
AUTHOR = "dh"
VERTXT = f"Version {VERSION} · {AUTHOR}"
LOGO = r"C:\Users\derDi\FBH_Auslegung_Excel\agn-logo.png"

FONT = "Arial"
NAME_ROW, SYM_ROW, UNIT_ROW = 4, 5, 6
R0, N_ROWS = 7, 200
R1 = R0 + N_ROWS - 1   # 206
EMU = 9525
# Graustufen-Palette (Schrift schwarz; Farbe nur Logo; Bewertung über Schriftfarbe grün/rot)
# BLUE bleibt als interne Spalten-Markierung (Eingabe vs. Berechnung) erhalten, wird aber nicht mehr als Textfarbe genutzt.
BLUE, BLACK, WHITE, GREY, NAVY = "0000FF", "000000", "FFFFFF", "808080", "000000"
HDR_FILL = PatternFill("solid", fgColor="D2D2D2")     # Tabellenkopf / Abschnittsbalken – einheitlich, kräftigeres Grau
SUB_FILL = PatternFill("solid", fgColor="E6E6E6")
INPUT_FILL = PatternFill("solid", fgColor="FFFFFF")   # EINGABE-/editierbare Zellen: weiß
CALC_FILL = PatternFill("solid", fgColor="ECECEC")    # BERECHNETE Zellen: leichtes Grau
ACCENT_FILL = PatternFill("solid", fgColor="E6E6E6")
RED_FILL = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")
GREEN_FILL = PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid")
EVAL_OK, EVAL_WARN = "1E8E1E", "D32F2F"   # Bewertung als Schriftfarbe (kräftiger): grün ok / rot Grenzwert
thin = Side(style="thin", color="C9C9C9")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
CEN = Alignment(horizontal="center")
LEFT = Alignment(horizontal="left")

# ---- Auslegungs-spezifisch (Haarlinien, Kopfunterstrich, Zebra) ----
INK = "000000"
GREY_HAIR = "C9C9C9"
_hair = Side(style="thin", color=GREY_HAIR)
_blk_med = Side(style="medium", color=INK)
HAIR_BORDER = Border(left=_hair, right=_hair, top=_hair, bottom=_hair)
HEAD_UNDERLINE = Border(left=_hair, right=_hair, top=_hair, bottom=_blk_med)   # schwarze Linie unter dem Kopf
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
ZEBRA_INPUT = PatternFill(start_color="FFF3F3F3", end_color="FFF3F3F3", fill_type="solid")   # Streifen weiße Eingabespalten
ZEBRA_CALC = PatternFill(start_color="FFE0E0E0", end_color="FFE0E0E0", fill_type="solid")    # Streifen graue Rechenspalten

def f(bold=False, color=BLACK, size=10, italic=False):
    return Font(name=FONT, bold=bold, color=color, size=size, italic=italic)

def fz(text, color=BLACK, bold=False, size=10, italic=False):
    """Formelzeichen mit echten Tiefstellungen: nach '_' folgt bis zum nächsten
    Leerzeichen ein tiefgestelltes Token (z. B. q_HL → q mit tiefem HL)."""
    if "_" not in text:
        return text
    sf = InlineFont(rFont=FONT, vertAlign="subscript", b=bold, i=italic, sz=size, color=color)
    parts, cur, i, n = [], "", 0, len(text)
    while i < n:
        if text[i] == "_" and i + 1 < n and not text[i + 1].isspace():
            if cur:
                parts.append(cur); cur = ""
            i += 1; tok = ""
            while i < n and not text[i].isspace():
                tok += text[i]; i += 1
            parts.append(TextBlock(sf, tok))
        else:
            cur += text[i]; i += 1
    if cur:
        parts.append(cur)
    return CellRichText(parts)

def col_px(ws, idx):
    dim = ws.column_dimensions[get_column_letter(idx)]
    if dim.hidden: return 0
    w = dim.width if dim.width else 8.43
    return int(round(w * 7)) + 5

def add_logo_corner(ws, last_col, height=40):
    if not os.path.exists(LOGO): return
    img = XLImage(LOGO)
    ratio = img.width / img.height
    img.height = height; img.width = int(height * ratio)
    right_edge = sum(col_px(ws, i) for i in range(1, last_col + 1))
    left = max(0, right_edge - img.width - 4)
    acc, c = 0, 1
    while c <= last_col:
        cw = col_px(ws, c)
        if cw and acc + cw > left: break
        acc += cw; c += 1
    marker = AnchorMarker(col=c - 1, colOff=(left - acc) * EMU, row=0, rowOff=1 * EMU)
    img.anchor = OneCellAnchor(_from=marker, ext=XDRPositiveSize2D(cx=img.width * EMU, cy=img.height * EMU))
    ws.add_image(img)

def disp_header(ws, title, last_col, bearb_cell=None, bearb_merge=None, project=True):
    ws.row_dimensions[1].height = 34
    t = ws["A1"]; t.value = title
    t.font = Font(name=FONT, bold=True, color=NAVY, size=15); t.alignment = Alignment(horizontal="left", vertical="center")
    add_logo_corner(ws, last_col, height=40)
    if project:
        a2 = ws["A2"]; a2.value = f'="Projekt-Nr.: "&{gPNr}'; a2.font = f(bold=True, color=NAVY, size=11); a2.alignment = LEFT
        a3 = ws["A3"]; a3.value = f'="Projekt: "&{gPName}'; a3.font = f(bold=True, color=NAVY, size=11); a3.alignment = LEFT
    if bearb_cell:
        if bearb_merge: ws.merge_cells(bearb_merge)
        b = ws[bearb_cell]; b.value = f'="Bearbeiter: "&{gBearb}'
        b.font = f(italic=True, color=GREY, size=9); b.alignment = Alignment(horizontal="right", vertical="center")

def setup_print(ws, area, titles=None, orientation="landscape", paper=9):
    # paper: 9 = A4, 8 = A3. Auf eine Seite breit skaliert (fit-to-width).
    ws.page_setup.orientation = orientation
    ws.page_setup.paperSize = paper
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_area = area
    if titles: ws.print_title_rows = titles
    ws.page_margins.left = ws.page_margins.right = 0.4
    ws.page_margins.top = ws.page_margins.bottom = 0.5
    ws.oddFooter.left.text = "FBH überschlägige Auslegung"
    ws.oddFooter.center.text = "Seite &P von &N"
    ws.oddFooter.right.text = "Stand &D"

def mini_header(ws, row, col, text):
    c = ws.cell(row=row, column=col, value=text)
    c.font = f(bold=True, color=BLACK); c.fill = HDR_FILL
    c.alignment = Alignment(wrap_text=True, horizontal="center"); c.border = BORDER

def two_col_table(ws, title_row, c0, title, hdr1, hdr2, rows, fmt2, n_empty=0):
    ws.cell(row=title_row, column=c0, value=title).font = f(bold=True, color=NAVY)
    ws.merge_cells(start_row=title_row, start_column=c0, end_row=title_row, end_column=c0 + 1)
    mini_header(ws, title_row + 1, c0, hdr1); mini_header(ws, title_row + 1, c0 + 1, hdr2)
    top = title_row + 2
    total = len(rows) + n_empty
    for i in range(total):
        r = top + i
        a = ws.cell(row=r, column=c0); b = ws.cell(row=r, column=c0 + 1)
        a.font = f(color=BLACK); a.fill = INPUT_FILL; a.border = BORDER
        b.font = f(color=BLACK); b.fill = INPUT_FILL; b.border = BORDER; b.number_format = fmt2; b.alignment = CEN
        if i < len(rows):
            a.value = rows[i][0]; b.value = rows[i][1]
            a.alignment = Alignment(horizontal="left" if isinstance(rows[i][0], str) else "center")
        else:
            a.alignment = LEFT
    return top, top + total - 1

wb = Workbook()

# ---- Referenzen auf KONSTANTEN ----
K = "'Konstanten'!"
FT_R = f"{K}$A$6:$A$10"; FT_AB = f"{K}$B$6:$B$10"; FT_AT = f"{K}$C$6:$C$10"   # EN-1264 Faktoren: R | a_B | a_T
RW_VAL_LIST = f"{K}$L$6:$L$25"    # Dropdown: R-Wert (Spalte L); Belag-Name steht daneben (Spalte K)
ZONE_RANGE = f"{K}$H$6:$I$13"; ZONE_LIST = f"{K}$H$6:$H$13"   # Zonen (G/H/I): Kürzel -> θF,max
PIPE_RANGE = f"{K}$A$18:$E$25"; PIPE_LIST = f"{K}$A$18:$A$25"

# Bodenbeläge (Name, R-Wert) – Quelle für die Konstanten-Tabelle UND den Dropdown-Hinweis
BELAEGE = [
    ("ohne Belag (roh)", 0.000), ("Keramik / Feinsteinzeug", 0.010), ("Fliesen 8 mm", 0.012),
    ("Naturstein / Marmor", 0.015), ("PVC / Vinyl (verklebt)", 0.020), ("Designboden (Klick)", 0.040),
    ("Linoleum 2,5 mm", 0.050), ("Klick-Vinyl (schwimmend)", 0.055), ("Laminat + Trittschall", 0.060),
    ("Parkett / Fertigparkett", 0.075), ("Mehrschichtparkett 14 mm", 0.090), ("Kork", 0.100),
    ("Teppich (dünn)", 0.100), ("Holzdielen", 0.110), ("Massivholzdielen 20 mm", 0.130),
    ("Teppich (mittel)", 0.140), ("Teppich (dick)", 0.170),
]

# =====================================================================
#  Blatt 1: GRUNDEINSTELLUNGEN  (Bezeichnungen ausgeschrieben)
# =====================================================================
g = wb.active
g.title = "Grundeinstellungen"
g.sheet_view.showGridLines = False
for col, w in (("A", 23), ("B", 15), ("C", 11), ("D", 8), ("E", 17), ("F", 7), ("G", 2),
               ("H", 23), ("I", 15), ("J", 14), ("K", 8), ("L", 17), ("M", 7)):
    g.column_dimensions[col].width = w
g.row_dimensions[1].height = 34
g["A1"].value = "Globale Grundeinstellungen – Fußbodenheizung (DIN EN 1264)"
g["A1"].font = Font(name=FONT, bold=True, color=NAVY, size=15); g["A1"].alignment = Alignment(horizontal="left", vertical="center")
add_logo_corner(g, 13, height=40)
for row, lab in ((2, "Projekt-Nr."), (3, "Projektname")):
    g.cell(row=row, column=1, value=lab).font = f(bold=True)
    g.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    v = g.cell(row=row, column=3, value=""); v.font = f(bold=True, color=BLACK); v.fill = INPUT_FILL; v.border = BORDER; v.alignment = LEFT
    g.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
g.cell(row=2, column=12, value="Bearbeiter:").font = f(italic=True, color=GREY, size=9)
g["L2"].alignment = Alignment(horizontal="right")
bf = g.cell(row=2, column=13, value=""); bf.font = f(bold=True, color=BLACK); bf.fill = INPUT_FILL; bf.border = BORDER; bf.alignment = CEN

def section(row, text, cL):
    for col in range(cL, cL + 6): g.cell(row=row, column=col).fill = HDR_FILL
    g.cell(row=row, column=cL, value=text).font = f(bold=True, color=BLACK)
def param(row, label, value, unit, note, fmt, cL):
    g.cell(row=row, column=cL, value=fz(label)).font = f()
    g.merge_cells(start_row=row, start_column=cL, end_row=row, end_column=cL + 1)
    v = g.cell(row=row, column=cL + 2, value=value); v.font = f(bold=True, color=BLACK); v.alignment = CEN; v.border = BORDER; v.fill = INPUT_FILL
    if fmt: v.number_format = fmt
    g.cell(row=row, column=cL + 3, value=unit).font = f(color=GREY)
    g.cell(row=row, column=cL + 4, value=note).font = f(italic=True, color=GREY, size=9)
    g.merge_cells(start_row=row, start_column=cL + 4, end_row=row, end_column=cL + 5)
def calcrow(row, label, formula, unit, fmt, cL, note=""):
    g.cell(row=row, column=cL, value=fz(label)).font = f(); g.merge_cells(start_row=row, start_column=cL, end_row=row, end_column=cL + 1)
    v = g.cell(row=row, column=cL + 2, value=formula); v.font = f(bold=True); v.alignment = CEN; v.border = BORDER; v.number_format = fmt; v.fill = CALC_FILL
    g.cell(row=row, column=cL + 3, value=unit).font = f(color=GREY)
    if note:
        g.cell(row=row, column=cL + 4, value=note).font = f(italic=True, color=GREY, size=9); g.merge_cells(start_row=row, start_column=cL + 4, end_row=row, end_column=cL + 5)

# ---- linke Spalte (Werte in C) ----
section(5, "Heizmedium / Temperaturen", 1)
param(6, "Vorlauftemperatur  θV", 35, "°C", "Auslegungs-Vorlauf", '0.0" °C"', 1)
param(7, "Rücklauftemperatur  θR", 28, "°C", "Auslegungs-Rücklauf", '0.0" °C"', 1)
calcrow(8, "Spreizung  θV − θR", "=C6-C7", "K", '0.0" K"', 1, "berechnet")
section(10, "Grenzwerte", 1)
param(11, "Maximale zulässige Kreislänge", 120, "m", "Obergrenze je Heizkreis", '0" m"', 1)
param(12, "Maximaler Druckverlust je Kreis", 15000, "Pa", "Warnschwelle", '#,##0" Pa"', 1)
section(14, "Stoffwerte Heizmedium (Wasser ~30 °C)", 1)
param(15, "Dichte  ρ", 995.7, "kg/m³", "Dichte des Mediums", '0.0', 1)
param(16, "spezifische Wärmekapazität  c_p", 4180, "J/(kg·K)", "spez. Wärmekapazität", '#,##0', 1)
param(17, "kinematische Viskosität  ν", 0.000000801, "m²/s", "für die Reynolds-Zahl", '0.00E+00', 1)
section(19, "Strömungsgrenzwerte", 1)
param(20, "Minimale Geschwindigkeit  v_min", 0.10, "m/s", "untere empfohlene Grenze", '0.00" m/s"', 1)
param(21, "Maximale Geschwindigkeit  v_max", 0.50, "m/s", "obere empfohlene Grenze", '0.00" m/s"', 1)
param(22, "Max. Volumenstrom je Kreis  V̇_max", 150, "l/h", "Warnschwelle je Heizkreis", '0" l/h"', 1)
section(23, "Druckverlust-Zuschläge", 1)
param(24, "Zuschlag Einzelwiderstände", 0.05, "-", "Formstücke (5 %)", '0%', 1)
param(25, "Aufschlag Verteiler je Kreis", 500, "Pa", "fester Wert je Verteiler", '#,##0" Pa"', 1)

# ---- rechte Spalte (Werte in J) ----
section(5, "Wärmeübergang / Fußbodenaufbau", 8)
param(6, "Wärmeübergang Oberfläche  α", 10.8, "W/m²K", "Boden → Raum (EN 1264)", '0.0', 8)
param(7, "Estrichüberdeckung über Rohr  s_ü", 0.045, "m", "Estrich über den Rohren", '0.000', 8)
param(8, "Wärmeleitfähigkeit Estrich  λ_E", 1.2, "W/mK", "Zementestrich typ. 1,2", '0.00', 8)
section(10, "Wärmeabgabe nach unten (Dämmung, global)", 8)
param(11, "Wärmedurchlasswiderstand  R_u", 2.0, "m²K/W", "Dämmung unter den Rohren", '0.00', 8)
param(12, "Temperatur unter der Decke  θ_u", 10, "°C", "Raum/Erdreich unter FBH", '0.0" °C"', 8)
calcrow(13, "Wärmestrom nach unten  q_u", "=((C6+C7)/2-J12)/J11", "W/m²", '0.0" W/m²"', 8, "Mittel θm = (θV+θR)/2")
section(15, "Rohrsystem-Auswahl (Werte aus Konstanten)", 8)
g.cell(row=16, column=8, value="Gewähltes Rohrsystem").font = f(); g.merge_cells("H16:I16")
v = g.cell(row=16, column=10, value="PE-Xa 17x2"); v.font = f(bold=True, color=BLACK); v.fill = INPUT_FILL; v.alignment = CEN; v.border = BORDER
g.merge_cells("J16:K16")
g.cell(row=16, column=12, value="Dropdown").font = f(italic=True, color=GREY, size=9); g.merge_cells("L16:M16")
calcrow(17, "→ Außendurchmesser  da", f"=VLOOKUP($J$16,{PIPE_RANGE},2,FALSE)", "mm", '0.0" mm"', 8, "aus Rohrbibliothek")
calcrow(18, "→ Wandstärke  s", f"=VLOOKUP($J$16,{PIPE_RANGE},3,FALSE)", "mm", '0.0" mm"', 8, "aus Rohrbibliothek")
calcrow(19, "→ Innendurchmesser  di", f"=VLOOKUP($J$16,{PIPE_RANGE},4,FALSE)", "mm", '0.0" mm"', 8, "di = da − 2·s")
calcrow(20, "→ Rauheit  k", f"=VLOOKUP($J$16,{PIPE_RANGE},5,FALSE)", "mm", '0.000" mm"', 8, "aus Rohrbibliothek")
calcrow(21, "→ Innendurchmesser  di", "=J19/1000", "m", '0.0000" m"', 8, "für die Berechnung")
calcrow(22, "→ Innenquerschnitt  A_i", "=PI()/4*J21^2", "m²", '0.000000" m²"', 8, "für die Berechnung")
gilt = g.cell(row=23, column=8, value="Gilt für Bauart A nach DIN EN 1264 (Heizrohre im Estrich).")
gilt.font = f(italic=True, color=GREY, size=9); g.merge_cells("H23:M23")
dv_pipe = DataValidation(type="list", formula1=f"={PIPE_LIST}", allow_blank=False)
g.add_data_validation(dv_pipe); dv_pipe.add(g["J16"])
setup_print(g, "A1:M25")

# ---- Globale Referenzen ----
G = "'Grundeinstellungen'!"
gV, gR = f"{G}$C$6", f"{G}$C$7"
gMax, gWarn = f"{G}$C$11", f"{G}$C$12"
grho, gcp, gnu = f"{G}$C$15", f"{G}$C$16", f"{G}$C$17"
gvmin, gvmax = f"{G}$C$20", f"{G}$C$21"
gVdot = f"{G}$C$22"
gzus, gVarm = f"{G}$C$24", f"{G}$C$25"
galp, gsu, glam = f"{G}$J$6", f"{G}$J$7", f"{G}$J$8"
gqdown = f"{G}$J$13"
gdim, gAi = f"{G}$J$21", f"{G}$J$22"
gPNr, gPName, gBearb = f"{G}$C$2", f"{G}$C$3", f"{G}$M$2"
# EN-1264-Parameter (Bauart A) – stehen jetzt im Blatt Konstanten (Werte in Spalte C)
gB = f"{K}$C$13"; gaue = f"{K}$C$14"; gaD = f"{K}$C$15"; gmue = f"{K}$C$16"; gmD = f"{K}$C$17"

def en_interp(Rcell, col_range):
    """Stückweise lineare Interpolation eines EN-1264-Faktors über R_λ,B (FT_R → col_range)."""
    m = f"MATCH({Rcell},{FT_R},1)"
    lo = f"INDEX({col_range},{m})"; hi = f"INDEX({col_range},{m}+1)"
    rlo = f"INDEX({FT_R},{m})"; rhi = f"INDEX({FT_R},{m}+1)"
    return f'IFERROR({lo}+({Rcell}-{rlo})/({rhi}-{rlo})*({hi}-{lo}),{lo})'

# =====================================================================
#  Blatt 2: AUSLEGUNG
# =====================================================================
rl = wb.create_sheet("Auslegung")
rl.sheet_view.showGridLines = False
# Überschriften vom Nutzer optimiert (Wortrennungen beibehalten); Formelzeichen ergänzt.
columns = [
    ("Heizkreis-\nverteiler", "HKV", "", 16, "text", BLUE),            # A
    ("Raum-Nr.", "", "", 10, "text", BLUE),                             # B
    ("Raum-\nbezeichnung", "", "", 18, "text", BLUE),                  # C
    ("Raum-fläche", "A_R", "[m²]", 10, '0.0" m²"', BLUE),              # D
    ("Raum-\ntemperatur", "θi", "[°C]", 11, '0.0" °C"', BLUE),         # E
    ("Heizlast", "Q", "[W]", 10, '#,##0" W"', BLUE),                   # F
    ("spez. Heizlast", "q_HL", "[W/m²]", 11, '0.0" W/m²"', BLACK),     # G  (=Q/Raumfläche)
    ("aktivier-\nbare Fläche", "A_F", "[m²]", 11, '0.0" m²"', BLUE),   # H
    ("R-Wert\nBodenbelag", "R_λ,B", "[m²·K/W]", 12, '0.000', BLUE),    # I (Dropdown R-Wert; Belag-Hinweis im Tooltip)
    ("Verlege–abstand", "VA", "[mm]", 11, '0" mm"', BLUE),             # J
    ("Anz.\nHK", "n", "[-]", 9, '0" HK"', BLUE),                       # K
    ("Zuleitungs-\nlänge", "L_zu", "[m]", 11, '0.0" m"', BLUE),        # L
    ("Zone", "", "", 7, "text", BLUE),                                 # M (Kürzel)
    ("log. Übertemperatur", "ΔθH", "[K]", 12, '0.00" K"', BLACK),      # N
    ("Bodenbelag-Faktor", "a_B", "[-]", 11, '0.000', BLACK),           # O  (EN 1264, interpoliert über R)
    ("Teilungs-\nfaktor", "a_T", "[-]", 11, '0.000', BLACK),           # P  (EN 1264, interpoliert über R)
    ("spez. Heiz-\nleistung", "q", "[W/m²]", 12, '0.0" W/m²"', BLACK), # Q  (EN-1264-Produktverfahren)
    ("Ober-\nflächen-\ntemperatur", "θF", "[°C]", 11, '0.0" °C"', BLACK),  # R
    ("max. Oberflächentemp.", "θF,max", "[°C]", 11, '0.0" °C"', BLACK),# S
    ("Leistung\nFBH", "Q_o", "[W]", 12, '#,##0" W"', BLACK),           # T
    ("Über-/Unter-\ndeckung", "ΔQ", "[W]", 12, '"+"#,##0" W";"-"#,##0" W";0" W"', BLACK),  # U
    ("Deckung", "", "[%]", 10, '0.0%', BLACK),                         # V
    ("Verlust nach unten", "Q_u", "[W]", 11, '#,##0" W"', BLACK),      # W
    ("maßgebl. Leistung", "Q_m", "[W]", 12, '#,##0" W"', BLACK),       # X  (=min(T;Q)+W)
    ("Rohrlänge Fläche", "L_F", "[m]", 11, '0.0" m"', BLACK),          # Y
    ("Rohrlänge Zuleitung", "L_Z", "[m]", 11, '0.0" m"', BLACK),       # Z
    ("Rohrlänge gesamt", "L_ges", "[m]", 11, '0.0" m"', BLACK),        # AA
    ("Rohrlänge je Kreis", "L_K", "[m]", 11, '0.0" m"', BLACK),        # AB
    ("Massenstrom je Kreis", "ṁ", "[kg/h]", 11, '0.0" kg/h"', BLACK),  # AC
    ("Volumen-strom je Kreis", "V̇", "[l/h]", 11, '0.0" l/h"', BLACK), # AD
    ("Geschwin-digkeit", "v", "[m/s]", 11, '0.000" m/s"', BLACK),      # AE
    ("Reynolds-Zahl", "Re", "[-]", 11, '#,##0', BLACK),                # AF
    ("Strömungsbereich", "", "[-]", 12, "text", BLACK),                # AG
    ("Rohrreibungszahl", "λ", "[-]", 11, '0.0000', BLACK),             # AH
    ("Δp Reibung", "Δp_R", "[Pa]", 11, '#,##0" Pa"', BLACK),           # AI
    ("spez. Druckverlust", "R", "[Pa/m]", 12, '#,##0" Pa/m"', BLACK),  # AJ
    ("Druck-\nverlust", "Δp", "[Pa]", 12, '#,##0" Pa"', BLACK),        # AK
]  # A..AK = 37
NCOL = len(columns)
LASTCOL = get_column_letter(NCOL)
INPUT_COLS = [j for j, c in enumerate(columns, start=1) if c[5] == BLUE]
CALC_COLS = [j for j, c in enumerate(columns, start=1) if c[5] == BLACK]
INPUT_STYLED = {7}   # spez. Heizlast: berechnet, aber optisch wie die Eingabespalten
for j, (name, sym, unit, width, fmt, color) in enumerate(columns, start=1):
    rl.column_dimensions[get_column_letter(j)].width = width
    is_inp = (color == BLUE) or (j in INPUT_STYLED)
    fill = HDR_FILL   # gesamter Tabellenkopf in kräftigerem Grau (einheitlich)
    for row, txt, sz in ((NAME_ROW, name, 10), (SYM_ROW, sym, 13), (UNIT_ROW, unit, 9)):
        bold = row != UNIT_ROW                          # Einheiten nicht fett, aber schwarz
        val = fz(txt, color=INK, bold=bold, size=sz) if row == SYM_ROW else txt
        c = rl.cell(row=row, column=j, value=val)
        c.fill = fill; c.font = f(bold=bold, color=INK, size=sz)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = HEAD_UNDERLINE if row == UNIT_ROW else HAIR_BORDER   # schwarze Linie unter dem Kopf
rl.row_dimensions[NAME_ROW].height = 34
rl.row_dimensions[SYM_ROW].height = 20
rl.row_dimensions[UNIT_ROW].height = 14

def F(r):
    return {
        7:  f'=IF($B{r}="","",IFERROR($F{r}/$D{r},""))',
        14: f'=IF($B{r}="","",IFERROR(({gV}-{gR})/LN(({gV}-$E{r})/({gR}-$E{r})),""))',
        15: f'=IF($B{r}="","",IFERROR({en_interp(f"$I{r}", FT_AB)},""))',
        16: f'=IF($B{r}="","",IFERROR({en_interp(f"$I{r}", FT_AT)},""))',
        17: (f'=IF(OR($B{r}="",$N{r}="",$O{r}="",$P{r}=""),"",'
             f'{gB}*$O{r}*$P{r}^(1-($J{r}/1000)/0.075)*{gaue}^{gmue}*{gaD}^{gmD}*$N{r})'),
        18: f'=IF(OR($B{r}="",$Q{r}=""),"",$E{r}+$Q{r}/{galp})',
        19: f'=IF($B{r}="","",IFERROR(VLOOKUP($M{r},{ZONE_RANGE},2,FALSE),29))',
        20: f'=IF(OR($B{r}="",$Q{r}=""),"",$Q{r}*$H{r})',
        21: f'=IF(OR($B{r}="",$T{r}=""),"",$T{r}-$F{r})',
        22: f'=IF(OR($B{r}="",$T{r}=""),"",IFERROR($T{r}/$F{r},""))',
        23: f'=IF(OR($B{r}="",$H{r}=""),"",{gqdown}*$H{r})',
        24: f'=IF(OR($B{r}="",$T{r}=""),"",MIN($T{r},$F{r})+$W{r})',
        25: f'=IF($B{r}="","",IFERROR($H{r}/($J{r}/1000),""))',
        26: f'=IF($B{r}="","",IFERROR(2*$L{r}*$K{r},""))',
        27: f'=IF(OR($B{r}="",$Y{r}="",$Z{r}=""),"",$Y{r}+$Z{r})',
        28: f'=IF(OR($B{r}="",$AA{r}=""),"",IFERROR($AA{r}/$K{r},""))',
        29: f'=IF(OR($B{r}="",$X{r}=""),"",IFERROR(($X{r}/$K{r})/({gcp}*({gV}-{gR}))*3600,""))',
        30: f'=IF(OR($B{r}="",$AC{r}=""),"",$AC{r}/{grho}*1000)',
        31: f'=IF(OR($B{r}="",$AC{r}=""),"",IFERROR($AC{r}/3600/{grho}/{gAi},""))',
        32: f'=IF(OR($B{r}="",$AE{r}=""),"",IFERROR($AE{r}*{gdim}/{gnu},""))',
        33: f'=IF(OR($B{r}="",$AF{r}=""),"",IF($AF{r}<2300,"laminar",IF($AF{r}<4000,"Übergangsbereich","turbulent")))',
        34: f'=IF(OR($B{r}="",$AF{r}=""),"",IF($AF{r}<2300,64/$AF{r},0.316/$AF{r}^0.25))',
        35: f'=IF(OR($B{r}="",$AH{r}="",$AB{r}=""),"",IFERROR($AH{r}*($AB{r}/{gdim})*({grho}/2)*$AE{r}^2,""))',
        36: f'=IF(OR($B{r}="",$AI{r}="",$AB{r}=""),"",IFERROR($AI{r}/$AB{r},""))',
        37: f'=IF(OR($B{r}="",$AI{r}=""),"",$AI{r}*(1+{gzus})+{gVarm})',
    }

examples = [   # Heizlast so gewählt, dass spez. Heizlast (Q/Raumfläche) <= 50 W/m², bunte Mischung
    ["HKV EG", "001", "Foyer/Empfang", 40, 20, 1800, 36, 0.00,  100, 3, 6,  "RZ"],   # 45 W/m²
    ["HKV EG", "002", "Großraumbüro",  60, 20, 2400, 55, 0.10,  150, 4, 8,  "AZ"],   # 40 W/m²
    ["HKV EG", "003", "Büro 1",        18, 20,  630, 16, 0.10,  150, 1, 5,  "AZ"],   # 35 W/m²
    ["HKV EG", "004", "Besprechung",   30, 20, 1500, 28, 0.075, 100, 2, 7,  "AZ"],   # 50 W/m²
    ["HKV EG", "005", "Flur",          25, 20,  750, 22, 0.05,  200, 1, 3,  "AZ"],   # 30 W/m²
    ["HKV OG", "101", "WC / Sanitär",  12, 24,  600,  9, 0.01,  100, 1, 10, "BAD"],  # 50 W/m²
    ["HKV OG", "102", "Teeküche",      10, 20,  450,  8, 0.01,  100, 1, 9,  "AZ"],   # 45 W/m²
    ["HKV OG", "103", "Archiv",        20, 18,  500, 18, 0.05,  200, 1, 12, "AZ"],   # 25 W/m²
    ["HKV OG", "104", "Technikraum",   15, 15,  300, 12, 0.00,  250, 1, 14, "AZ"],   # 20 W/m²
]
for idx, r in enumerate(range(R0, R1 + 1)):
    fr = F(r)
    for ji, j in enumerate(INPUT_COLS):
        c = rl.cell(row=r, column=j)
        if idx < len(examples): c.value = examples[idx][ji]
        c.font = f(color=INK); c.fill = INPUT_FILL; c.border = HAIR_BORDER   # Eingabe: weiß
        fmt = columns[j - 1][4]
        if fmt != "text": c.number_format = fmt
        c.alignment = Alignment(horizontal="left" if j in (1, 2, 3) else "center")
    for j in CALC_COLS:
        c = rl.cell(row=r, column=j, value=fr[j])
        c.font = f(color=INK); c.fill = CALC_FILL; c.border = HAIR_BORDER   # Berechnet: leichtes Grau
        fmt = columns[j - 1][4]
        if fmt != "text": c.number_format = fmt
        c.alignment = CEN

rl.freeze_panes = "C7"
dv_zone = DataValidation(type="list", formula1=f"={ZONE_LIST}", allow_blank=True, showErrorMessage=False)
rl.add_data_validation(dv_zone); dv_zone.add(f"M{R0}:M{R1}")
dv_rw = DataValidation(type="list", formula1=f"={RW_VAL_LIST}", allow_blank=True, showErrorMessage=False)
rl.add_data_validation(dv_rw); dv_rw.add(f"I{R0}:I{R1}")

# Bewertung über die SCHRIFTFARBE: leicht grün = ok, leicht rot = Grenzwert überschritten.
# (Keine farbigen Zellhintergründe; Werte bleiben sonst schwarz.)
def cf_font(col, ok_formula, bad_formula):
    rng = f"{col}{R0}:{col}{R1}"
    rl.conditional_formatting.add(rng, FormulaRule(formula=[bad_formula], font=Font(color=EVAL_WARN)))
    rl.conditional_formatting.add(rng, FormulaRule(formula=[ok_formula], font=Font(color=EVAL_OK)))
cf_font("R", f'AND($R{R0}<>"",$R{R0}<=$S{R0})', f'AND($R{R0}<>"",$R{R0}>$S{R0})')
cf_font("T", f'AND($T{R0}<>"",$F{R0}<>"",$T{R0}>=$F{R0})', f'AND($T{R0}<>"",$F{R0}<>"",$T{R0}<$F{R0})')
cf_font("V", f'AND($V{R0}<>"",$V{R0}>=1)', f'AND($V{R0}<>"",$V{R0}<1)')
cf_font("AB", f'AND($AB{R0}<>"",$AB{R0}<={gMax})', f'AND($AB{R0}<>"",$AB{R0}>{gMax})')
cf_font("AD", f'AND($AD{R0}<>"",$AD{R0}<={gVdot})', f'AND($AD{R0}<>"",$AD{R0}>{gVdot})')
cf_font("AK", f'AND($AK{R0}<>"",$AK{R0}<={gWarn})', f'AND($AK{R0}<>"",$AK{R0}>{gWarn})')
cf_font("AE", f'AND($AE{R0}<>"",$AE{R0}>={gvmin},$AE{R0}<={gvmax})',
        f'AND($AE{R0}<>"",OR($AE{R0}<{gvmin},$AE{R0}>{gvmax}))')
rl.conditional_formatting.add(f"H{R0}:H{R1}", FormulaRule(
    formula=[f'AND($H{R0}<>"",$D{R0}<>"",$H{R0}>$D{R0})'], font=Font(color=EVAL_WARN)))
# Zebra-Streifen (filterfest via SUBTOTAL: zählt nur sichtbare Zeilen) für die Lesbarkeit.
_zform = f'AND($B{R0}<>"",MOD(SUBTOTAL(103,$B${R0}:$B{R0}),2)=0)'
_inp_z = sorted(INPUT_COLS)                # weiße Eingabespalten
_calc_z = sorted(CALC_COLS)                # graue Rechenspalten (inkl. Bewertungs-/rot-grün-Spalten)
_zrange = lambda cols: " ".join(f"{get_column_letter(j)}{R0}:{get_column_letter(j)}{R1}" for j in cols)
rl.conditional_formatting.add(_zrange(_inp_z), FormulaRule(formula=[_zform], fill=ZEBRA_INPUT))
rl.conditional_formatting.add(_zrange(_calc_z), FormulaRule(formula=[_zform], fill=ZEBRA_CALC))
# schlanker Standard: Zwischen-/Sekundärspalten ausblenden (jederzeit einblendbar)
for col in ["N", "O", "P", "S", "U", "W", "X", "Y", "Z", "AA", "AC", "AF", "AG", "AH", "AI", "AJ"]:
    rl.column_dimensions[col].hidden = True
# sichtbare Spalten schmaler, damit das Blatt auf A4-Querformat passt
narrow = {"A": 13, "B": 11, "C": 16, "D": 8, "E": 11, "F": 9, "G": 10, "H": 9, "I": 11,
          "J": 9, "K": 8, "L": 11, "M": 6, "Q": 11, "R": 11, "T": 9, "V": 8,
          "AB": 9, "AD": 10, "AE": 10, "AK": 10}
for col, w in narrow.items():
    rl.column_dimensions[col].width = w
rl.row_dimensions[NAME_ROW].height = 46   # mehr Höhe, da schmale Spalten stärker umbrechen
# ---- Kopf (Graustufen, Schrift schwarz; Farbe nur über das Logo) ----
rl.row_dimensions[1].height = 32
t = rl.cell(row=1, column=1, value="Fußbodenheizung – Auslegung (DIN EN 1264)")
t.font = Font(name=FONT, bold=True, color=INK, size=16)
t.alignment = Alignment(horizontal="left", vertical="center")
add_logo_corner(rl, NCOL, height=36)
meta = rl.cell(row=2, column=1, value=f'="Projekt-Nr.  "&{gPNr}&"        Projekt  "&{gPName}')
meta.font = f(color=INK, size=11); meta.alignment = LEFT
rl.merge_cells(f"AB2:{LASTCOL}2")
bb = rl.cell(row=2, column=28, value=f'="Bearbeiter  "&{gBearb}')
bb.font = f(italic=True, color=INK, size=9)
bb.alignment = Alignment(horizontal="right", vertical="center")
_hdr_line = Border(top=Side(style="thin", color=INK))   # dünne schwarze Briefkopflinie
for col in range(1, NCOL + 1):
    rl.cell(row=3, column=col).border = _hdr_line
rl.row_dimensions[3].height = 8
rl.auto_filter.ref = f"A{UNIT_ROW}:{LASTCOL}{R1}"   # Filterfunktion auf der Einheitenzeile
setup_print(rl, f"A1:{LASTCOL}{R1}", titles="1:6", orientation="landscape", paper=9)
rl.page_margins.left = rl.page_margins.right = 0.2   # schmalere Ränder → mehr Platz auf A4
rl.page_margins.top = rl.page_margins.bottom = 0.3

# =====================================================================
#  Blatt 3: KONTROLLE
# =====================================================================
ov = wb.create_sheet("Kontrolle")
ov.sheet_view.showGridLines = False
for col, w in (("A", 44), ("B", 16), ("C", 11), ("D", 40)):
    ov.column_dimensions[col].width = w
AUS = "'Auslegung'!"
def ovrow(r, label, formula, unit, fmt='#,##0', note=""):
    ov.cell(row=r, column=1, value=label).font = f()
    c = ov.cell(row=r, column=2, value=formula); c.font = f(bold=True); c.number_format = fmt
    c.alignment = CEN; c.border = BORDER; c.fill = CALC_FILL   # berechnet: grau
    ov.cell(row=r, column=3, value=unit).font = f(color=GREY)
    if note: ov.cell(row=r, column=4, value=note).font = f(italic=True, color=GREY, size=9)
c = ov.cell(row=5, column=1, value="Summen / Kennzahlen"); c.font = f(bold=True, color=BLACK); c.fill = HDR_FILL
for col in (2, 3, 4): ov.cell(row=5, column=col).fill = HDR_FILL
ovrow(6, "Anzahl Räume", f"=COUNTA({AUS}B{R0}:B{R1})", "-", '0')
ovrow(7, "Gesamte Heizlast", f"=SUM({AUS}F{R0}:F{R1})", "W")
ovrow(8, "Gesamte FBH-Leistung (nach oben)", f"=SUM({AUS}T{R0}:T{R1})", "W")
ovrow(9, "Verlustleistung nach unten", f"=SUM({AUS}W{R0}:W{R1})", "W")
ovrow(10, "Maßgebliche Gesamtleistung (Hydraulik)", f"=SUM({AUS}X{R0}:X{R1})", "W")
ovrow(11, "Deckungsgrad gesamt", "=IFERROR(B8/B7,0)", "%", '0.0%')
ovrow(12, "Gesamte Rohrlänge", f"=SUM({AUS}AA{R0}:AA{R1})", "m")
ovrow(13, "Anzahl Heizkreise gesamt", f"=SUM({AUS}K{R0}:K{R1})", "-", '0')
ovrow(14, "Gesamtmassenstrom (Pumpe)", f"=SUM({AUS}AC{R0}:AC{R1})", "kg/h", '#,##0')
ovrow(15, "Gesamtvolumenstrom (Pumpe)", f"=SUM({AUS}AD{R0}:AD{R1})", "l/h", '#,##0')
ovrow(16, "Max. Druckverlust eines Kreises", f"=IFERROR(MAX({AUS}AK{R0}:AK{R1}),0)", "Pa")
ovrow(17, "Max. Strömungsgeschwindigkeit", f"=IFERROR(MAX({AUS}AE{R0}:AE{R1}),0)", "m/s", '0.000')
c = ov.cell(row=19, column=1, value="Warnungen"); c.font = f(bold=True, color=BLACK); c.fill = HDR_FILL
for col in (2, 3, 4): ov.cell(row=19, column=col).fill = HDR_FILL
ovrow(20, "Räume unterdeckt", f'=COUNTIF({AUS}V{R0}:V{R1},"<1")', "-", '0', "Heizlast nicht gedeckt")
ovrow(21, "Kreise zu lang", f'=COUNTIF({AUS}AB{R0}:AB{R1},">"&{gMax})', "-", '0', "> max. Kreislänge")
ovrow(22, "Druckverlust über Warnschwelle", f'=COUNTIF({AUS}AK{R0}:AK{R1},">"&{gWarn})', "-", '0')
ovrow(23, "Oberflächentemperatur zu hoch", f'=SUMPRODUCT(({AUS}R{R0}:R{R1}<>"")*({AUS}R{R0}:R{R1}>{AUS}S{R0}:S{R1}))', "-", '0')
ovrow(24, "Geschwindigkeit außerhalb v_min–v_max", f'=SUMPRODUCT(({AUS}AE{R0}:AE{R1}<>"")*(({AUS}AE{R0}:AE{R1}<{gvmin})+({AUS}AE{R0}:AE{R1}>{gvmax})))', "-", '0', "Kreise")
ovrow(25, "Volumenstrom je Kreis zu hoch", f'=COUNTIF({AUS}AD{R0}:AD{R1},">"&{gVdot})', "-", '0', "> max. Volumenstrom")
for r in (20, 21, 22, 23, 24, 25):
    ov.conditional_formatting.add(f"B{r}", FormulaRule(formula=[f"$B${r}>0"], font=Font(color=EVAL_WARN)))
    ov.conditional_formatting.add(f"B{r}", FormulaRule(formula=[f"$B${r}=0"], font=Font(color=EVAL_OK)))
disp_header(ov, "Kontrolle / Auswertung", 4, "D2")
setup_print(ov, "A1:D25")

# =====================================================================
#  Blatt 4: HEIZKREISVERTEILER  (+ Massenstrom)
# =====================================================================
hv = wb.create_sheet("HKV")
hv.sheet_view.showGridLines = False
for col, w in (("A", 20), ("B", 10), ("C", 10), ("D", 17), ("E", 14), ("F", 14), ("G", 16), ("H", 50)):
    hv.column_dimensions[col].width = w
for j, h in enumerate(["Heizkreisverteiler (HKV)", "Anzahl Kreise", "Spreizung [K]", "maßgebl. Leistung [W]",
                       "Massenstrom [kg/h]", "Volumenstrom [l/h]", "max. Druckverlust [Pa]",
                       "angebundene Räume (Raum-Nr.)"], start=1):
    c = hv.cell(row=4, column=j, value=h)
    c.font = f(bold=True, color=BLACK); c.fill = HDR_FILL
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); c.border = BORDER
hv.row_dimensions[4].height = 30
Aaus = f"{AUS}$A${R0}:$A${R1}"
N_HKV = 50
for i in range(N_HKV):
    r = 5 + i
    af = (f'=IFERROR(INDEX({Aaus},MATCH(0,COUNTIF($A$4:$A{r-1},{Aaus})+IF({Aaus}="",1,0),0)),"")')
    cell = hv.cell(row=r, column=1); cell.value = ArrayFormula(f"A{r}", af); cell.font = f(); cell.alignment = LEFT; cell.border = BORDER; cell.fill = CALC_FILL
    b = hv.cell(row=r, column=2, value=f'=IF($A{r}="","",SUMIF({Aaus},$A{r},{AUS}$K${R0}:$K${R1}))')
    sp = hv.cell(row=r, column=3, value=f"=IF($A{r}=\"\",\"\",{gV}-{gR})")   # globale Spreizung (überall gleich)
    c = hv.cell(row=r, column=4, value=f'=IF($A{r}="","",SUMIF({Aaus},$A{r},{AUS}$X${R0}:$X${R1}))')
    d = hv.cell(row=r, column=5, value=f'=IF($A{r}="","",SUMIF({Aaus},$A{r},{AUS}$AC${R0}:$AC${R1}))')
    e = hv.cell(row=r, column=6, value=f'=IF($A{r}="","",SUMIF({Aaus},$A{r},{AUS}$AD${R0}:$AD${R1}))')
    p = hv.cell(row=r, column=7)
    p.value = ArrayFormula(f"G{r}", f'=IF($A{r}="","",MAX(IF({Aaus}=$A{r},{AUS}$AK${R0}:$AK${R1})))')
    for cc, fmt in ((b, '0'), (sp, '0.0" K"'), (c, '#,##0" W"'), (d, '#,##0" kg/h"'), (e, '#,##0" l/h"'), (p, '#,##0" Pa"')):
        cc.font = f(); cc.alignment = CEN; cc.number_format = fmt; cc.border = BORDER; cc.fill = CALC_FILL
    rm = hv.cell(row=r, column=8)
    rm.value = ArrayFormula(f"H{r}", f'=IF($A{r}="","",_xlfn.TEXTJOIN(", ",TRUE,IF({Aaus}=$A{r},{AUS}$B${R0}:$B${R1},"")))')
    rm.font = f(); rm.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True); rm.border = BORDER; rm.fill = CALC_FILL
hv.cell(row=5 + N_HKV + 1, column=1,
        value="Baut sich automatisch aus der Auslegung auf (Array-Formeln, Excel & LibreOffice).").font = f(italic=True, color=GREY, size=9)
disp_header(hv, "Heizkreisverteiler – Übersicht je HKV", 8, "F2", "F2:G2")
setup_print(hv, f"A1:H{5 + N_HKV + 1}")

# =====================================================================
#  Blatt 5: VERIFIKATION
# =====================================================================
vf = wb.create_sheet("Verifikation")
vf.sheet_view.showGridLines = False
for col, w in (("A", 13), ("B", 7), ("C", 7), ("D", 11), ("E", 11), ("F", 13),
               ("G", 13), ("H", 12), ("I", 10), ("J", 9), ("K", 10), ("L", 8)):
    vf.column_dimensions[col].width = w
# EN-1264-Produktverfahren – KEIN Systemfaktor. Vergleich berechnet vs. Hersteller (Wärmestrom nach oben).
mh = vf.cell(row=5, column=1, value="Berechnung nach DIN EN 1264 (Bauart A): je Auslegungspunkt wird die Wärmestromdichte q aus B, a_B, a_T, a_ü, a_D und ΔθH ermittelt und mit dem Hersteller-/Referenzwert verglichen.")
mh.font = f(bold=True, color=NAVY, size=10); vf.merge_cells("A5:L5")
for cc in range(1, 13): vf.cell(row=5, column=cc).fill = ACCENT_FILL
vf["A5"].alignment = Alignment(horizontal="left", vertical="center")
vf.row_dimensions[5].height = 18
vf.cell(row=6, column=1, value="Verglichen wird die spezifische Heizleistung NACH OBEN (Wärmestrom zur Raumseite). Die Wärmeabgabe nach unten ist hier bewusst NICHT enthalten – wie in den Hersteller-/EN-1264-Kennwerten.").font = f(italic=True, color=GREY, size=9)
vf.merge_cells("A6:L6")
vhdr = ["Verlege-\nabstand [mm]", "θV\n[°C]", "θR\n[°C]", "θi\n[°C]", "R-Wert\n[m²K/W]",
        "Hersteller q\nnach oben\n[W/m²]", "q berechnet\nEN 1264\n[W/m²]", "Abweichung\n[W/m²]", "Abweichung\n[%]",
        "ΔθH\n[K]", "a_B\n[-]", "a_T\n[-]"]
VT = 8
for j, h in enumerate(vhdr, start=1):
    c = vf.cell(row=VT, column=j, value=h)
    c.fill = HDR_FILL                       # Kopf einheitlich grau (egal ob Eingabe- oder Rechenspalte)
    c.font = f(bold=True, color=BLACK)       # keine weiße Schrift
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); c.border = BORDER
vf.row_dimensions[VT].height = 48
V0 = VT + 1
# Beispiel-Auslegungspunkte: [VA, θV, θR, θi, R, Hersteller-q]
# Reale Herstellerdaten: REHAU 17×2-System, Estrichüberdeckung s_ü = 45 mm (= globaler Default).
vpts = [
    [300, 38, 30, 21, 0.125, 30],
    [250, 38, 30, 21, 0.125, 33],
    [200, 38, 30, 21, 0.125, 36],
    [150, 38, 30, 21, 0.125, 40],
    [100, 38, 30, 21, 0.125, 44],
    [50,  38, 30, 21, 0.125, 49],
    [300, 45, 35, 15, 0.050, 74],
    [250, 45, 35, 15, 0.050, 83],
    [200, 45, 35, 15, 0.050, 94],
    [150, 45, 35, 15, 0.050, 105],
    [100, 45, 35, 15, 0.050, 119],
    [50,  45, 35, 15, 0.050, 135],
]
V1 = V0 + len(vpts) - 1
for i, pt in enumerate(vpts):
    r = V0 + i
    inputs = {1: (pt[0], '0" mm"'), 2: (pt[1], '0.0" °C"'), 3: (pt[2], '0.0" °C"'),
              4: (pt[3], '0.0" °C"'), 5: (pt[4], '0.000'), 6: (pt[5], '0.0" W/m²"')}
    for col, (val, fmt) in inputs.items():
        cc = vf.cell(row=r, column=col, value=val); cc.font = f(color=BLACK); cc.fill = INPUT_FILL
        cc.alignment = CEN; cc.number_format = fmt; cc.border = BORDER
    mT = f"(1-($A{r}/1000)/0.075)"
    forms = {
        10: (f'=IF($A{r}="","",IFERROR(($B{r}-$C{r})/LN(($B{r}-$D{r})/($C{r}-$D{r})),""))', '0.00" K"'),
        11: (f'=IF($A{r}="","",IFERROR({en_interp(f"$E{r}", FT_AB)},""))', '0.0000'),
        12: (f'=IF($A{r}="","",IFERROR({en_interp(f"$E{r}", FT_AT)},""))', '0.0000'),
        7:  (f'=IF(OR($A{r}="",$J{r}="",$K{r}="",$L{r}=""),"",'
             f'{gB}*$K{r}*$L{r}^{mT}*{gaue}^{gmue}*{gaD}^{gmD}*$J{r})', '0.0" W/m²"'),
        8:  (f'=IF(OR($A{r}="",$F{r}=""),"",$F{r}-$G{r})', '0.0" W/m²"'),
        9:  (f'=IF(OR($A{r}="",$F{r}=""),"",IFERROR(($F{r}-$G{r})/$G{r},""))', '0.0%'),
    }
    for col, (formula, fmt) in forms.items():
        cc = vf.cell(row=r, column=col, value=formula); cc.font = f(); cc.fill = CALC_FILL   # berechnet: grau
        cc.alignment = CEN; cc.number_format = fmt; cc.border = BORDER
vf.conditional_formatting.add(f"I{V0}:I{V1}", FormulaRule(formula=[f'AND($I{V0}<>"",ABS($I{V0})>0.05)'], font=Font(color=EVAL_WARN)))
vf.conditional_formatting.add(f"I{V0}:I{V1}", FormulaRule(formula=[f'AND($I{V0}<>"",ABS($I{V0})<=0.05)'], font=Font(color=EVAL_OK)))
# Summenzeile: Gesamtabweichung über alle Punkte (zeigt die Güte des EN-1264-Modells)
VS = V1 + 1
slab = vf.cell(row=VS, column=1, value="Gesamtabweichung (Σ über alle Punkte):")
slab.font = f(bold=True, color=NAVY); slab.alignment = Alignment(horizontal="right", vertical="center")
vf.merge_cells(start_row=VS, start_column=1, end_row=VS, end_column=7)
for cc in range(1, 8): vf.cell(row=VS, column=cc).fill = SUB_FILL
sh = vf.cell(row=VS, column=8, value=f"=SUM($H${V0}:$H${V1})")
sh.font = f(bold=True); sh.number_format = '"+"0.0" W/m²";"-"0.0" W/m²";0" W/m²"'
sh.alignment = CEN; sh.border = BORDER; sh.fill = SUB_FILL
si = vf.cell(row=VS, column=9, value=f'=IFERROR(SUM($H${V0}:$H${V1})/SUM($G${V0}:$G${V1}),"")')
si.font = f(bold=True); si.number_format = '"+"0.0%;"-"0.0%;0.0%'
si.alignment = CEN; si.border = BORDER
vf.conditional_formatting.add(f"I{VS}", FormulaRule(formula=[f'AND($I{VS}<>"",ABS($I{VS})>0.05)'], font=Font(color=EVAL_WARN)))
vf.conditional_formatting.add(f"I{VS}", FormulaRule(formula=[f'AND($I{VS}<>"",ABS($I{VS})<=0.05)'], font=Font(color=EVAL_OK)))
vf.row_dimensions[VS].height = 18
nr = VS + 2
for i, (txt, kind) in enumerate([
    ("Hinweise", "h"),
    ("• Jede Zeile vergleicht einen eigenen Auslegungspunkt (Spalten A–E) mit dem Hersteller-/Referenzwert (Spalte F).", ""),
    ("• q berechnet (Spalte G): vollständiges EN-1264-Produktverfahren (Bauart A), direkt aus den Faktoren B, a_B, a_T, a_ü, a_D.", ""),
    ("• Hersteller-q (Spalte F): vorbelegt mit 12 Punkten des Rehau 17×2-Systems (s_ü = 45 mm, Verlegeabstand 50–300 mm).", ""),
    ("• Die Faktoren a_B, a_T je R-Wert sowie B, a_ü, a_D stehen im Blatt Konstanten und bestimmen das Ergebnis.", ""),
    ("• Andere nach EN 1264 zertifizierte Systeme (Uponor, Purmo, Buderus, Kermi …) liegen nah an diesen Werten.", "i")]):
    c = vf.cell(row=nr + i, column=1, value=txt)
    if kind == "h": c.font = f(bold=True, color=NAVY, size=11)
    elif kind == "i": c.font = f(italic=True, color=GREY, size=9)
    else: c.font = f()
disp_header(vf, "Verifikation – EN-1264-Berechnung vs. Herstellerdaten", 12, project=True)
setup_print(vf, f"A1:L{max(nr + 8, 22)}")

# =====================================================================
#  Blatt 6: METHODIK
# =====================================================================
m = wb.create_sheet("Methodik")
m.sheet_view.showGridLines = False
for col, w in (("A", 3), ("B", 46), ("C", 3), ("D", 48), ("E", 3), ("F", 48)):
    m.column_dimensions[col].width = w

def methcol(col, start, items):
    rr = start
    for text, kind in items:
        c = m.cell(row=rr, column=col, value=text)
        if kind == "h": c.font = Font(name=FONT, bold=True, color=NAVY, size=11)
        elif kind == "fb": c.font = Font(name=FONT, bold=True, color=BLACK, size=10)
        elif kind == "i": c.font = Font(name=FONT, italic=True, color=GREY, size=9)
        else: c.font = Font(name=FONT, size=10)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        rr += 1
    return rr

m["B4"].value = "Formelzeichen / Variablen"; m["B4"].font = Font(name=FONT, bold=True, color=NAVY, size=11)
m["D4"].value = "Berechnungsschritte"; m["D4"].font = Font(name=FONT, bold=True, color=NAVY, size=11)
m.merge_cells("D4:F4")
vars_ = [
    ("θV, θR — Vor-/Rücklauftemperatur [°C]", ""), ("θi — Raumtemperatur [°C]", ""),
    ("θ_u — Temperatur unter der Decke [°C]", ""), ("ΔθH — log. Übertemperatur [K]", ""),
    ("q — spez. Heizleistung nach oben [W/m²]", ""), ("α — Wärmeübergang Oberfläche [W/m²K]", ""),
    ("R_Belag — Widerstand Bodenbelag [m²K/W]", ""), ("s_ü — Estrichüberdeckung [m]", ""),
    ("λ_E — Wärmeleitfähigkeit Estrich [W/mK]", ""), ("B — Systemkoeffizient (EN 1264) [W/m²K]", ""),
    ("a_B, a_T — Bodenbelag-/Teilungsfaktor [-]", ""), ("a_ü, a_D — Überdeckungs-/Durchmesserfaktor [-]", ""),
    ("R_u — Widerstand nach unten [m²K/W]", ""), ("q_u — Wärmestrom nach unten [W/m²]", ""),
    ("θF, θF,max — Oberflächentemp. / Grenze [°C]", ""), ("di, A_i — Rohr-Innen-Ø / -Querschnitt", ""),
    ("ρ, c_p, ν — Dichte / Wärmekap. / Viskosität", ""), ("ṁ, V̇ — Massen-/Volumenstrom [kg/h]/[l/h]", ""),
    ("v, Re, λ — Geschw. / Reynolds / Reibung", ""), ("L_Kreis — Rohrlänge je Kreis [m]", ""),
    ("Δp — Druckverlust [Pa]", ""),
]
steps_l = [
    ("1)  ΔθH = (θV−θR) / ln((θV−θi)/(θR−θi))", "fb"),
    ("Mittlere treibende Temperaturdifferenz Heizwasser ↔ Raum (log. Mittel).", "i"), ("", ""),
    ("2)  m_T = 1 − Abstand/0,075 ; m_ü = 100·(0,045 − s_ü) ; m_D = 250·(D − 0,020)", "fb"),
    ("Exponenten aus Verlegeabstand, Estrichüberdeckung und Rohr-Außendurchmesser.", "i"), ("", ""),
    ("3)  q = B · a_B · a_T^m_T · a_ü^m_ü · a_D^m_D · ΔθH", "fb"),
    ("EN-1264-Produktverfahren (Bauart A); a_B, a_T je R-Wert aus den Konstanten.", "i"), ("", ""),
    ("4)  θF = θi + q/α", "fb"),
    ("Bodenoberflächentemperatur; Prüfung gegen θF,max.", "i"), ("", ""),
    ("5)  Leistung oben = q · Fläche ; Deckung % = Leistung / Heizlast", "fb"),
    ("Prüft, ob die FBH die Heizlast deckt (≥ 100 %).", "i"),
]
steps_r = [
    ("6)  q_u = ((θV+θR)/2 − θ_u)/R_u ; Verlust = q_u · Fläche", "fb"),
    ("Wärmeabgabe nach unten durch die Dämmung.", "i"), ("", ""),
    ("7)  maßgebl. Leistung = MIN(Leistung ; Heizlast) + Verlust", "fb"),
    ("Tatsächlich transportierte Wärme (Basis für Massenstrom).", "i"), ("", ""),
    ("8)  L_Kreis = (Fläche/Verlegeabstand + 2·Zuleitung·Kreise)/Kreise", "fb"),
    ("Rohrlänge je Kreis; Prüfung gegen max. Kreislänge.", "i"), ("", ""),
    ("9)  ṁ = (maßgebl. Leistung/Kreise)/(c_p·(θV−θR))", "fb"),
    ("v = V̇/A_i ; Re = v·di/ν ; Re < 2300 = laminar.", "i"), ("", ""),
    ("10) Δp = λ·(L/di)·(ρ/2)·v²·(1+Zuschlag) + Verteiler", "fb"),
    ("Druckverlust je Kreis; Kontrolle gegen Warnschwelle.", "i"), ("", ""),
    ("Hinweise", "h"),
    ("• Gilt für Bauart A nach DIN EN 1264 (Heizrohre im Estrich).", ""),
    ("• Faktoren a_B, a_T, a_ü, a_D und B stehen editierbar im Blatt Konstanten.", ""),
    ("• Wärmeabgabe nach unten global; pauschale Δp-Zuschläge (Hydraulik wie Basis-Tool).", ""),
]
end_b = methcol(2, 5, [(fz(t), k) for t, k in vars_])
methcol(4, 6, steps_l)
methcol(6, 6, steps_r)
disp_header(m, "Methodik & Annahmen", 6, project=False)
setup_print(m, f"A1:F{end_b}")

# =====================================================================
#  Blatt 7: KONSTANTEN
# =====================================================================
kt = wb.create_sheet("Konstanten")
kt.sheet_view.showGridLines = False
# Spalten F und J sind WEISSE Leerspalten zwischen den Tabellen (optische Trennung).
for col, w in (("A", 15), ("B", 11), ("C", 11), ("D", 9), ("E", 9), ("F", 3),
               ("G", 16), ("H", 8), ("I", 11), ("J", 3), ("K", 24), ("L", 11)):
    kt.column_dimensions[col].width = w
kt.cell(row=3, column=1, value="Nachschlage-Tabellen (editierbar). Tabellen durch weiße Zellen getrennt.").font = f(italic=True, color=GREY, size=9)

# --- Tabelle 1: EN-1264-Faktoren (Bauart A) | Spalten A–C ---
kt.cell(row=4, column=1, value="EN-1264-Faktoren (Bauart A)").font = f(bold=True, color=NAVY)
kt.merge_cells("A4:C4")
mini_header(kt, 5, 1, "R-Wert [m²K/W]"); mini_header(kt, 5, 2, "a_B [-]"); mini_header(kt, 5, 3, "a_T [-]")
en_fac = [(0.00, 1.0846, 1.2393), (0.05, 0.7880, 1.1957), (0.10, 0.6143, 1.1646),
          (0.15, 0.5067, 1.1403), (0.20, 0.4311, 1.1160)]
for i, (Rv, aBv, aTv) in enumerate(en_fac):
    rr = 6 + i
    for col, val, fmt in ((1, Rv, '0.000'), (2, aBv, '0.0000'), (3, aTv, '0.0000')):
        c = kt.cell(row=rr, column=col, value=val); c.font = f(color=BLACK); c.fill = INPUT_FILL
        c.border = BORDER; c.alignment = CEN; c.number_format = fmt
# (Zeile 11 bleibt weiß = Trennung)

# --- Tabelle 2: EN-1264-Parameter | Spalten A–C (Werte in C) ---
kt.cell(row=12, column=1, value="EN-1264-Parameter").font = f(bold=True, color=NAVY)
kt.merge_cells("A12:C12")
GE = "'Grundeinstellungen'!"
en_par = [
    ("Systemkoeffizient  B", 6.7, "W/m²K", '0.0', False),
    ("Überdeckungsfaktor  a_ü", 1.07, "-", '0.000', False),
    ("Rohrdurchmesser-Faktor  a_D", 1.06, "-", '0.000', False),
    ("Überdeckungs-Exponent  m_ü", f"=100*(0.045-{GE}$J$7)", "-", '0.000', True),
    ("Durchmesser-Exponent  m_D", f"=250*({GE}$J$17/1000-0.020)", "-", '0.000', True),
]
for i, (lab, val, unit, fmt, calc) in enumerate(en_par):
    r = 13 + i
    lc = kt.cell(row=r, column=1, value=fz(lab)); lc.font = f(); kt.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    vc = kt.cell(row=r, column=3, value=val); vc.alignment = CEN; vc.border = BORDER; vc.number_format = fmt
    vc.font = f(bold=True, color=BLACK)
    vc.fill = CALC_FILL if calc else INPUT_FILL   # berechnet grau / Eingabe weiß
    kt.cell(row=r, column=4, value=unit).font = f(color=GREY)
# (Zeile 18 bleibt weiß = Trennung)

# --- Tabelle 3: Rohrbibliothek | Spalten A–E ---
kt.cell(row=19, column=1, value="Rohrbibliothek (di = da − 2·s)").font = f(bold=True, color=NAVY)
for j, h in enumerate(["Rohrsystem", "da [mm]", "s [mm]", "di [mm]", "k [mm]"], start=1):
    mini_header(kt, 20, j, h)
pipes = [("PE-Xa 17x2", 17, 2.0, 0.007), ("PE-Xa 16x2", 16, 2.0, 0.007),
         ("PE-Xa 20x2", 20, 2.0, 0.007), ("PE-RT 14x2", 14, 2.0, 0.007)]
colmap = {1: 0, 2: 1, 3: 2, 5: 3}
for i in range(8):
    r = 21 + i
    for col in range(1, 6):
        cell = kt.cell(row=r, column=col); cell.border = BORDER
        if col == 4:
            cell.value = f'=IF(B{r}="","",B{r}-2*C{r})'; cell.font = f(color=BLACK); cell.number_format = '0.0'; cell.alignment = CEN; cell.fill = CALC_FILL
        else:
            cell.font = f(color=BLACK); cell.fill = INPUT_FILL
            cell.alignment = Alignment(horizontal="left" if col == 1 else "center")
            if col in (2, 3): cell.number_format = '0.0'
            if col == 5: cell.number_format = '0.000'
            if i < len(pipes): cell.value = pipes[i][colmap[col]]

# --- Tabelle 4: Zonen | Spalten G–I (Spalte F weiß) ---
kt.cell(row=4, column=7, value="Zonen / max. Oberflächentemp.").font = f(bold=True, color=NAVY)
kt.merge_cells("G4:I4")
mini_header(kt, 5, 7, "Zone"); mini_header(kt, 5, 8, "Kürzel"); mini_header(kt, 5, 9, "θF,max [°C]")
zdata = [("Aufenthaltszone", "AZ", 29), ("Randzone", "RZ", 35), ("Badezimmer", "BAD", 33)]
for i in range(8):
    r = 6 + i
    for col in (7, 8, 9):
        cell = kt.cell(row=r, column=col); cell.font = f(color=BLACK); cell.fill = INPUT_FILL; cell.border = BORDER
        cell.alignment = Alignment(horizontal="left" if col <= 8 else "center")
        if col == 9: cell.number_format = '0" °C"'
        if i < len(zdata): cell.value = zdata[i][col - 7]

# --- Tabelle 5: R-Werte Bodenbeläge | Spalten K–L (Spalte J weiß) ---
two_col_table(kt, 4, 11, "R-Werte Bodenbeläge", "Bodenbelag", "R [m²·K/W]", BELAEGE, '0.000', n_empty=3)

disp_header(kt, "Konstanten / Bibliotheken", 12, project=False)
setup_print(kt, "A1:L28")
kt.page_setup.fitToHeight = 1   # garantiert eine A4-Quer-Seite

# =====================================================================
#  Blatt 8: CHANGELOG
# =====================================================================
cl = wb.create_sheet("Changelog")
cl.sheet_view.showGridLines = False
for col, w in (("A", 12), ("B", 14), ("C", 115)):
    cl.column_dimensions[col].width = w
for j, h in enumerate(["Version", "Datum", "Änderungen"], start=1):
    c = cl.cell(row=5, column=j, value=h)
    c.font = f(bold=True, color=BLACK); c.fill = HDR_FILL
    c.alignment = Alignment(horizontal="left", vertical="center"); c.border = BORDER
r = 6
for ver, datum, changes in reversed(CHANGELOG):
    start = r
    for ch in changes:
        cl.cell(row=r, column=3, value="• " + ch).font = f()
        cl.cell(row=r, column=3).alignment = Alignment(wrap_text=True, vertical="top")
        for col in (1, 2, 3): cl.cell(row=r, column=col).border = BORDER
        r += 1
    a = cl.cell(row=start, column=1, value=ver); a.font = f(bold=True); a.alignment = Alignment(vertical="top")
    b = cl.cell(row=start, column=2, value=datum); b.font = f(); b.alignment = Alignment(vertical="top")
    if r - start > 1:
        cl.merge_cells(start_row=start, start_column=1, end_row=r - 1, end_column=1)
        cl.merge_cells(start_row=start, start_column=2, end_row=r - 1, end_column=2)
disp_header(cl, "Changelog – FBH-Auslegung", 3, project=False)
cl.cell(row=3, column=3, value=VERTXT).font = f(italic=True, color=GREY, size=9)
cl.cell(row=3, column=3).alignment = Alignment(horizontal="right")
setup_print(cl, f"A1:C{r-1}")

# =====================================================================
#  Blatt 0: DECKBLATT  (Übersichtsdiagramm aus Zellen – in Excel editierbar)
# =====================================================================
db = wb.create_sheet("Deckblatt", 0)
db.sheet_view.showGridLines = False
for col, w in (("A", 14), ("B", 14), ("C", 14), ("D", 5), ("E", 16), ("F", 16),
               ("G", 16), ("H", 5), ("I", 14), ("J", 14), ("K", 14)):
    db.column_dimensions[col].width = w
RED_S = PatternFill("solid", fgColor="D2D2D2")     # Graustufen: alle drei Bänder gleiches Grau
GREEN_S = PatternFill("solid", fgColor="D2D2D2")
BODY_S = PatternFill("solid", fgColor="F7F7F7")
AMB_H = PatternFill("solid", fgColor="FFE699")
AMB_B = PatternFill("solid", fgColor="FBF3D9")
AMBER = "B88600"

def box(r1, c1, r2, c2, value=None, fill=None, font=None, align="left", valign="top", border=True):
    db.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cc = db.cell(row=r, column=c)
            if fill: cc.fill = fill
            if border: cc.border = BORDER
    tl = db.cell(row=r1, column=c1)
    if value is not None: tl.value = value
    if font: tl.font = font
    tl.alignment = Alignment(horizontal=align, vertical=valign, wrap_text=True)

db.row_dimensions[1].height = 28
db.merge_cells("A1:H1")
db["A1"].value = "Fußbodenheizung – Auslegung nach DIN EN 1264"
db["A1"].font = Font(name=FONT, bold=True, color=NAVY, size=18); db["A1"].alignment = Alignment(vertical="center")
db.merge_cells("A2:H2"); db["A2"].value = "Eingaben · Verarbeitung · Ergebnisse"; db["A2"].font = f(color=GREY, size=12)
db.merge_cells("A3:K3")
db["A3"].value = f'="Projekt-Nr.: "&{gPNr}&"          Projekt: "&{gPName}&"          Bearbeiter: "&{gBearb}'
db["A3"].font = f(bold=True, color=NAVY, size=11)
add_logo_corner(db, 11, height=34)

# Spalten-Köpfe (fett)
box(5, 1, 5, 3, "EINGABEN", fill=HDR_FILL, font=f(bold=True, color=BLACK, size=13), align="center", valign="center")
box(5, 5, 5, 7, "BERECHNUNG", fill=RED_S, font=f(bold=True, color=BLACK, size=13), align="center", valign="center")
box(5, 9, 5, 11, "ERGEBNISSE", fill=GREEN_S, font=f(bold=True, color=BLACK, size=13), align="center", valign="center")

# Formelzeichen mit echten Tief-/Hochstellungen (für die Berechnungs-Spalte):
# code 0 = normal, 1 = tiefgestellt, 2 = hochgestellt (Exponent)
def rsub(*segs):
    sub = InlineFont(rFont=FONT, vertAlign="subscript", b=True, sz=10, color=NAVY)
    sup = InlineFont(rFont=FONT, vertAlign="superscript", b=True, sz=10, color=NAVY)
    out = []
    for t, code in segs:
        if code == 1: out.append(TextBlock(sub, t))
        elif code == 2: out.append(TextBlock(sup, t))
        else: out.append(t)
    return CellRichText(out)

# Eine Information = eine Zelle (keine zusammenhängenden Mehrzeilen-Blöcke, keine Fehl-Umbrüche)
def db_item(r, c1, c2, text, kind):
    if kind == "h":        # Zwischenüberschrift
        box(r, c1, r, c2, text, fill=SUB_FILL, font=f(bold=True, color=NAVY, size=10.5), align="left", valign="center")
    elif kind == "fmla":   # Formel (Rich-Text mit Tiefstellung) – klar sichtbar in der Berechnungs-Spalte
        box(r, c1, r, c2, text, fill=None, font=f(bold=True, color=NAVY, size=10.5), align="left", valign="center")
    elif kind == "b":      # Aufzählungspunkt
        box(r, c1, r, c2, "•  " + text, fill=BODY_S, font=f(color=BLACK, size=10), align="left", valign="center")
    else:                  # Leerzeile (Auffüllung für gleiche Spaltenhöhe)
        box(r, c1, r, c2, "", fill=BODY_S, align="left", valign="center")

def db_col(c1, c2, items, start=6):
    for i, (text, kind) in enumerate(items):
        db_item(start + i, c1, c2, text, kind)

eingaben = [
    ("Globale Vorgaben", "h"),
    ("Vorlauf- und Rücklauftemperatur", "b"),
    ("Rohrsystem und Stoffwerte", "b"),
    ("Wärmeübergang und Estrichaufbau", "b"),
    ("Dämmung nach unten", "b"),
    ("Grenzwerte (Länge, Druck, Strömung)", "b"),
    ("Je Raum", "h"),
    ("Raumfläche und aktivierbare Fläche", "b"),
    ("Heizlast und Raumtemperatur", "b"),
    ("Bodenbelag-Widerstand, Verlegeabstand", "b"),
    ("Anzahl Heizkreise, Zuleitung, Zone", "b"),
    ("Heizkreisverteiler", "b"),
    ("EN-1264-Faktoren", "h"),
    ("B, a_B, a_T, a_ü, a_D (Konstanten)", "b"),
    ("", ""),
]
berechnung = [
    ("Heizleistung nach oben (EN 1264)", "h"),
    (rsub(("q = B · a", 0), ("B", 1), (" · a", 0), ("T", 1), ("mT", 2), (" · a", 0), ("ü", 1), ("mü", 2),
          (" · a", 0), ("D", 1), ("mD", 2), (" · Δθ", 0), ("H", 1)), "fmla"),
    (rsub(("Δθ", 0), ("H", 1), (" = (θ", 0), ("V", 1), (" − θ", 0), ("R", 1),
          (") / ln[(θ", 0), ("V", 1), (" − θ", 0), ("i", 1), (") / (θ", 0),
          ("R", 1), (" − θ", 0), ("i", 1), (")]", 0)), "fmla"),
    (rsub(("m", 0), ("T", 1), (" = 1 − Abstand / 0,075", 0)), "fmla"),
    ("Faktoren je R-Wert (Konstanten-Tabelle)", "b"),
    ("Oberflächentemperatur", "h"),
    (rsub(("θ", 0), ("F", 1), (" = θ", 0), ("i", 1), (" + q / α", 0)), "fmla"),
    ("Hydraulik / Druckverlust", "h"),
    ("Maßgebliche Leistung → Massenstrom", "b"),
    (rsub(("Δp = λ · (L / d", False), ("i", True), (") · (ρ / 2) · v²", False)), "fmla"),
    ("Strömung: Reynolds-Zahl, Rohrreibung", "b"),
    ("Rohrlängen je Heizkreis", "b"),
    ("Prüfungen (Ampelbewertung)", "h"),
    ("Deckung, Temperatur, Druckverlust", "b"),
    ("", ""),
]
ergebnisse = [
    ("Je Raum", "h"),
    ("Spezifische Heizleistung", "b"),
    ("Oberflächentemperatur", "b"),
    ("Deckungsgrad (Ampelbewertung)", "b"),
    ("Rohrlänge je Heizkreis (Ampel)", "b"),
    ("Massenstrom und Volumenstrom", "b"),
    ("Strömungsgeschwindigkeit", "b"),
    ("Druckverlust je Heizkreis (Ampel)", "b"),
    ("Je Heizkreisverteiler", "h"),
    ("Leistung und Volumenstrom", "b"),
    ("Maximaler Druckverlust", "b"),
    ("Gesamt", "h"),
    ("Kontrolle und Warnübersicht", "b"),
    ("", ""),
    ("", ""),
]
db_col(1, 3, eingaben)
db_col(5, 7, berechnung)
db_col(9, 11, ergebnisse)
NROW = max(len(eingaben), len(berechnung), len(ergebnisse))
last = 6 + NROW - 1
# Pfeile zwischen den Spalten
box(6, 4, last, 4, "→", font=f(bold=True, color="777777", size=22), align="center", valign="center", border=False)
box(6, 8, last, 8, "→", font=f(bold=True, color="777777", size=22), align="center", valign="center", border=False)
db.row_dimensions[5].height = 20
for r in range(6, last + 1): db.row_dimensions[r].height = 18

# Fußzeile: Formelzeichen-Legende + Verweis auf Methodik
box(last + 2, 1, last + 2, 11,
    "Formelzeichen:  θV / θR  Vorlauf- / Rücklauftemperatur · θi  Raumtemperatur · θF  Oberflächentemperatur · "
    "B  Systemkoeffizient · a_B / a_T  Bodenbelag- / Teilungsfaktor · a_ü / a_D  Überdeckungs- / Durchmesserfaktor · "
    "α  Wärmeübergangskoeffizient · λ  Rohrreibungszahl · ρ  Dichte · v  Strömungsgeschwindigkeit",
    fill=None, font=f(italic=True, color=GREY, size=9), align="left", valign="center", border=False)
box(last + 3, 1, last + 3, 11,
    "Überschlägige Auslegung – Vereinfachungen und Annahmen siehe Blatt „Methodik“.",
    fill=None, font=f(italic=True, color=GREY, size=9), align="left", valign="center", border=False)
db.page_setup.orientation = "landscape"
db.page_setup.paperSize = 9
db.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
db.page_setup.fitToWidth = 1; db.page_setup.fitToHeight = 1
db.print_area = f"A1:K{last + 3}"
db.page_margins.left = db.page_margins.right = 0.3
db.page_margins.top = db.page_margins.bottom = 0.3

# =====================================================================
#  Blatt 1: ANLEITUNG  (zweites Blatt)
# =====================================================================
an = wb.create_sheet("Anleitung", 1)
an.sheet_view.showGridLines = False
for col, w in (("A", 2), ("B", 60), ("C", 3), ("D", 60)):
    an.column_dimensions[col].width = w

def anl(col, start, items):
    rr = start
    for text, kind in items:
        c = an.cell(row=rr, column=col, value=text)
        if kind == "h": c.font = Font(name=FONT, bold=True, color=NAVY, size=12)
        elif kind == "sh": c.font = Font(name=FONT, bold=True, color=BLACK, size=10.5)
        elif kind == "i": c.font = Font(name=FONT, italic=True, color=GREY, size=9)
        else: c.font = Font(name=FONT, size=10)
        c.alignment = Alignment(vertical="top")
        rr += 1
    return rr

left = [
    ("Aufbau der Mappe", "h"),
    ("• Deckblatt – Übersicht (Input / Verarbeitung / Output)", ""),
    ("• Grundeinstellungen – globale Werte", ""),
    ("• Auslegung – Raumliste mit allen Berechnungen", ""),
    ("• Kontrolle – Summen & Warnungen", ""),
    ("• HKV – Auswertung je Heizkreisverteiler", ""),
    ("• Verifikation – Abgleich mit Herstellerdaten", ""),
    ("• Anleitung – diese Seite", ""),
    ("• Methodik – Formeln & Annahmen", ""),
    ("• Konstanten – Nachschlagetabellen", ""),
    ("• Changelog – Versionen", ""),
    ("", ""),
    ("Farblegende", "h"),
    ("• Blau (gelb hinterlegt) = Eingabe", ""),
    ("• Schwarz = berechnet (nicht ändern)", ""),
    ("• Grün = OK,  Rot = Warnung / Grenzwert überschritten", ""),
    ("• Nur die blauen Eingabezellen bearbeiten.", ""),
    ("", ""),
    ("Ampel-Legende (warum ist eine Zelle rot?)", "h"),
    ("• Deckung < 100 % → Heizlast nicht gedeckt", ""),
    ("• Oberflächentemperatur > Zonen-Grenztemperatur", ""),
    ("• Rohrlänge je Kreis > max. Kreislänge", ""),
    ("• Druckverlust je Kreis > Warnschwelle", ""),
    ("• Geschwindigkeit außerhalb der Grenzen", ""),
    ("• Volumenstrom je Kreis > Grenzwert", ""),
    ("• aktivierbare Fläche > Raumfläche", ""),
    ("Grenzwerte stehen in den Grundeinstellungen,", "i"),
    ("Zonen-Grenztemperaturen in den Konstanten.", "i"),
]
right = [
    ("Schritt für Schritt", "h"),
    ("1. Grundeinstellungen ausfüllen: Temperaturen, Rohr-", ""),
    ("    system, Stoffwerte, Grenzwerte.", ""),
    ("2. Konstanten prüfen / erweitern (Rohre, R-Werte, Zonen).", ""),
    ("3. Verifikation: Systemfaktor f kalibrieren – VOR der", ""),
    ("    Raumauslegung (er wirkt auf alle Räume).", ""),
    ("4. Bauteilliste (Schedule) aus Revit nach Excel exportieren", ""),
    ("    – Spalten A–F (HKV, Raum-Nr., Bezeichnung, Raum-", ""),
    ("    fläche, Raumtemperatur, Heizlast).", ""),
    ("5. Export in die Auslegung übernehmen (A–F).", ""),
    ("6. Spalte G (spez. Heizlast) frei lassen – wird berechnet.", ""),
    ("7. Spalten H–M je Raum ergänzen (aktiv. Fläche, R-Wert,", ""),
    ("    Verlegeabstand, Kreise, Zuleitung, Zone).", ""),
    ("8. Ergebnisse prüfen: Ampel in Auslegung + 'Kontrolle'.", ""),
    ("", ""),
    ("Spalten der Auslegung", "h"),
    ("Aus Revit (A–F): HKV, Raum-Nr., Bezeichnung,", ""),
    ("   Raumfläche, Raumtemperatur, Heizlast.", ""),
    ("Berechnet (G): spez. Heizlast (nicht ändern).", ""),
    ("Im Tool (H–M): aktiv. Fläche, R-Wert, Verlege-", ""),
    ("   abstand, Kreise, Zuleitung, Zone.", ""),
    ("", ""),
    ("Mehrere Verteiler / Geschosse auslegen", "h"),
    ("• Verteiler je Raum in Spalte A (HKV) eintragen", ""),
    ("   (z. B. HKV EG, HKV OG, HKV 1.OG …).", ""),
    ("• Das Blatt 'HKV' wertet jeden Verteiler automatisch", ""),
    ("   aus (Kreise, Leistung, Volumenstrom, max. Δp).", ""),
    ("• Beliebig viele Verteiler in EINER Liste – einfach", ""),
    ("   weitere Räume ergänzen.", ""),
]
end_l = anl(2, 3, left)
end_r = anl(4, 3, right)
disp_header(an, "Anleitung / Bedienung", 4, project=False)
setup_print(an, f"A1:D{max(end_l, end_r)}")

# Blattreihenfolge: Anleitung nach hinten, direkt vor die Methodik
SHEET_ORDER = ["Deckblatt", "Grundeinstellungen", "Auslegung", "Kontrolle", "HKV",
               "Verifikation", "Anleitung", "Methodik", "Konstanten", "Changelog"]
wb._sheets.sort(key=lambda ws: SHEET_ORDER.index(ws.title))
wb.active = 0

# =====================================================================
# Blattschutz: nur Formelzellen sperren (Eingabezellen bleiben editierbar).
# Schutz ohne Passwort → kann bei Bedarf über 'Überprüfen ▸ Blattschutz aufheben' gelöst werden.
def protect_formulas(ws):
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            v = cell.value
            is_formula = isinstance(v, ArrayFormula) or (isinstance(v, str) and v.startswith("="))
            cell.protection = Protection(locked=is_formula)
    ws.protection.sheet = True
    # Zeilenhöhe/Spaltenbreite trotz Blattschutz änderbar lassen (z. B. HKV-Raumliste umbrechen)
    ws.protection.formatRows = False
    ws.protection.formatColumns = False
for ws in wb.worksheets:
    protect_formulas(ws)
# Auslegung: Filtern trotz Blattschutz zulassen (Sortieren bleibt durch gesperrte Formeln eingeschränkt)
rl.protection.autoFilter = False
rl.protection.sort = False

# =====================================================================
wb.properties.version = VERSION
wb.properties.creator = AUTHOR
wb.properties.lastModifiedBy = AUTHOR
wb.properties.keywords = f"FBH-Auslegung v{VERSION} ({AUTHOR})"
try:
    wb.calculation.fullCalcOnLoad = True
except Exception:
    pass
OUT = r"C:\Users\derDi\FBH_Auslegung_Excel\FBH_Auslegung_EN1264.xlsx"
wb.save(OUT)
print("gespeichert:", OUT, "| Version", VERSION, AUTHOR)
